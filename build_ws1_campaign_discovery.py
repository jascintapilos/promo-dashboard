from __future__ import annotations

import json
import os
from datetime import date, timedelta
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.comments import Comment
from openpyxl.formatting.rule import CellIsRule, FormulaRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.utils import get_column_letter


OUT = Path.cwd() / "WS1_Campaign_Discovery_and_Gap_Analysis_v0.1.xlsx"
SAMPLE_LABEL = "SAMPLE DATA – REPLACE"
MAX_DATA_ROW = 306

# restrained presentation palette
NAVY = "172033"
NAVY_2 = "24314D"
GOLD = "BF9440"
CREAM = "F7F0DC"
INPUT = "FFF9E8"
CALC = "EAF1F7"
WHITE = "FFFFFF"
TEXT = "252A34"
MUTED = "6F7787"
GRID = "D9DEE7"
RED = "F9DEDC"
RED_TEXT = "A33A32"
AMBER = "FFF0CE"
AMBER_TEXT = "8A5E12"
GREEN = "DFF0E5"
GREEN_TEXT = "28623D"
BLUE = "E4EEF7"
EXTERNAL = "FCE7C8"

thin = Side(style="thin", color=GRID)
border_bottom = Border(bottom=thin)


def set_title(ws, title, subtitle, last_col):
    ws.sheet_view.showGridLines = False
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=last_col)
    ws.cell(1, 1, title)
    ws.cell(1, 1).font = Font(name="Aptos Display", size=20, bold=True, color=WHITE)
    ws.cell(1, 1).fill = PatternFill("solid", fgColor=NAVY)
    ws.cell(1, 1).alignment = Alignment(vertical="center")
    ws.row_dimensions[1].height = 34
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=last_col)
    ws.cell(2, 1, subtitle)
    ws.cell(2, 1).font = Font(name="Aptos", size=10, italic=True, color=TEXT)
    ws.cell(2, 1).fill = PatternFill("solid", fgColor=CREAM)
    ws.cell(2, 1).alignment = Alignment(vertical="center", wrap_text=True)
    ws.row_dimensions[2].height = 30


def style_header_row(ws, row, last_col):
    for c in range(1, last_col + 1):
        cell = ws.cell(row, c)
        cell.fill = PatternFill("solid", fgColor=NAVY)
        cell.font = Font(name="Aptos", size=9, bold=True, color=WHITE)
        cell.alignment = Alignment(vertical="center", wrap_text=True)
        cell.border = Border(bottom=Side(style="medium", color=GOLD))
    ws.row_dimensions[row].height = 42


def style_table_body(ws, start_row, end_row, start_col, end_col, calculated_cols=None, external_cols=None):
    calculated_cols = set(calculated_cols or [])
    external_cols = set(external_cols or [])
    for row in ws.iter_rows(min_row=start_row, max_row=end_row, min_col=start_col, max_col=end_col):
        for cell in row:
            if cell.column in calculated_cols:
                color = CALC
            elif cell.column in external_cols:
                color = EXTERNAL
            else:
                color = INPUT
            cell.fill = PatternFill("solid", fgColor=color)
            cell.font = Font(name="Aptos", size=9, color=TEXT)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = border_bottom


def add_table(ws, ref, name, style="TableStyleMedium2"):
    tab = Table(displayName=name, ref=ref)
    tab.tableStyleInfo = TableStyleInfo(name=style, showFirstColumn=False, showLastColumn=False, showRowStripes=True, showColumnStripes=False)
    ws.add_table(tab)


def add_list_validation(ws, cell_range, defined_name):
    dv = DataValidation(type="list", formula1=f"={defined_name}", allow_blank=True)
    dv.error = "Select a value from the dropdown list."
    dv.errorTitle = "Invalid selection"
    dv.prompt = "Choose from the approved list."
    dv.promptTitle = "Select value"
    ws.add_data_validation(dv)
    dv.add(cell_range)
    return dv


def define_name(wb, name, sheet, col_letter, start, end):
    dn = DefinedName(name, attr_text=f"'{sheet}'!${col_letter}${start}:${col_letter}${end}")
    wb.defined_names.add(dn)


def section(ws, row, title, start_col, end_col):
    ws.merge_cells(start_row=row, start_column=start_col, end_row=row, end_column=end_col)
    c = ws.cell(row, start_col, title)
    c.fill = PatternFill("solid", fgColor=GOLD)
    c.font = Font(name="Aptos", size=10, bold=True, color=WHITE)
    c.alignment = Alignment(vertical="center")
    ws.row_dimensions[row].height = 22


def card(ws, label_cell, value_cell, label, formula, number_format="0"):
    ws[label_cell] = label
    ws[label_cell].font = Font(name="Aptos", size=9, bold=True, color=TEXT)
    ws[label_cell].fill = PatternFill("solid", fgColor="ECEEF2")
    ws[label_cell].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws[value_cell] = formula
    ws[value_cell].font = Font(name="Aptos Display", size=17, bold=True, color=NAVY)
    ws[value_cell].fill = PatternFill("solid", fgColor=WHITE)
    ws[value_cell].alignment = Alignment(horizontal="center", vertical="center")
    ws[value_cell].number_format = number_format


wb = Workbook()
default = wb.active
wb.remove(default)
sheet_names = ["README", "Campaign Inventory", "Gap Summary", "Segment Map", "Data Source Map", "Action Tracker", "Lists"]
for s in sheet_names:
    wb.create_sheet(s)

readme = wb["README"]
inventory = wb["Campaign Inventory"]
gap = wb["Gap Summary"]
segment = wb["Segment Map"]
data_map = wb["Data Source Map"]
actions = wb["Action Tracker"]
lists = wb["Lists"]

# -----------------------------------------------------------------------------
# Lists and named ranges
# -----------------------------------------------------------------------------
list_values = {
    "Brands": ["WS1"],
    "CampaignSources": ["Stakeholder", "Sales", "VIP VM", "CK Analytics CRM", "Promotion", "Management", "Other"],
    "CampaignCategories": ["Acquisition", "Retention", "Reactivation", "Cross-Product", "VIP Progression", "Seasonal", "Payday", "Product Launch", "Standard Recurring Bonus", "General Promotion", "Other"],
    "CampaignStatuses": ["Draft", "Scheduled", "Live", "Completed", "Paused", "Cancelled", "Sunset", "Unknown"],
    "Frequencies": ["One-Time", "Weekly", "Biweekly", "Monthly", "Payday", "Seasonal", "Always-On", "Ad Hoc", "Unknown"],
    "ObjectiveClarityStatuses": ["Clear", "Partly Clear", "Unclear", "Not Provided"],
    "SegmentClarityStatuses": ["Clear", "Partly Clear", "Too Broad", "Unclear", "Not Provided"],
    "DataAvailabilityStatuses": ["Available", "Partial", "Pending", "Not Available", "Unknown"],
    "GapSeverities": ["Low", "Medium", "High", "Critical", "Pending Review"],
    "ClarificationStatuses": ["Not Started", "Requested", "In Progress", "Received", "Blocked", "Closed"],
    "RecommendedActions": ["Confirm Objective", "Define Segment Criteria", "Request Result Data", "Confirm Campaign Owner", "Review Mechanic", "Review Bonus Cap", "Review Turnover", "Merge Similar Campaigns", "Run Controlled Test", "Prepare Performance Analysis", "Continue Monitoring", "Consider Sunset", "No Immediate Action"],
    "DeliveryChannels": ["SMS", "MT", "Website Dialog", "Telegram", "WhatsApp", "Email", "Banner", "Popup", "VIP Direct Contact", "Multi-Channel", "Other"],
    "QCStatuses": ["Not Checked", "Passed", "Passed with Conditions", "Failed", "Recheck Required", "Unknown"],
    "OwnerGroups": ["Promotion", "Sales", "VIP", "Stakeholder", "CK Analytics", "Management", "Finance", "Unknown"],
    "Priorities": ["Low", "Medium", "High", "Critical"],
    "ActionStatuses": ["Not Started", "In Progress", "Waiting for Response", "Blocked", "Completed", "Cancelled"],
    "BonusTypes": ["Deposit Bonus", "Free Credit", "Cashback", "Free Spin", "Reload Bonus", "Voucher", "Mystery Reward", "No Bonus", "Other"],
    "Currencies": ["MYR", "SGD", "USD", "THB", "Rp"],
    "Products": ["Sports", "Casino", "Slots", "Live Casino", "Lottery", "Poker", "Multi-Product", "Unknown"],
    "VIPTiers": ["Non-VIP", "Bronze", "Silver", "Gold", "Platinum", "Diamond", "Multiple", "Unknown"],
    "PlayerActivityStatuses": ["New", "Active", "At Risk", "Inactive", "Lapsed", "Reactivated", "Unknown"],
    "YesNoUnknown": ["Yes", "No", "Unknown"],
    "RecurringStatuses": ["Recurring", "One-Time", "Unknown"],
    "ObjectiveSources": ["Campaign Brief", "Stakeholder Confirmation", "Historical Practice", "Inferred", "Not Documented"],
    "MechanicGapStatuses": ["None", "Minor", "Review Required", "Major", "Unknown"],
    "MeasurementStatuses": ["Ready", "Partial", "Not Ready", "Blocked"],
    "DataQualityStatuses": ["Good", "Acceptable", "Needs Review", "Poor", "Unknown"],
    "DataMapStatuses": ["Available", "Partial", "Pending", "Not Available", "Blocked", "Unknown"],
    "AccessStatuses": ["Yes", "Partial", "No", "Unknown"],
    "OverlapRisks": ["None", "Low", "Medium", "High", "Unknown"],
    "FilterAllBrands": ["All", "WS1"],
}

lists.sheet_view.showGridLines = False
lists["A1"] = "List Name"
lists["A1"].font = Font(bold=True, color=WHITE)
lists["A1"].fill = PatternFill("solid", fgColor=NAVY)
for idx, (name, vals) in enumerate(list_values.items(), start=1):
    col = get_column_letter(idx)
    lists.cell(1, idx, name)
    lists.cell(1, idx).font = Font(bold=True, color=WHITE)
    lists.cell(1, idx).fill = PatternFill("solid", fgColor=NAVY)
    for r, value in enumerate(vals, start=2):
        lists.cell(r, idx, value)
    define_name(wb, name, "Lists", col, 2, len(vals) + 1)
    lists.column_dimensions[col].width = max(18, min(32, max(len(str(v)) for v in vals) + 3))

threshold_col = len(list_values) + 2
tc = get_column_letter(threshold_col)
lists.cell(1, threshold_col, "Readiness Threshold")
lists.cell(2, threshold_col, "Ready for Analysis")
lists.cell(3, threshold_col, "Partially Ready")
lists.cell(1, threshold_col + 1, "Value")
lists.cell(2, threshold_col + 1, 0.80)
lists.cell(3, threshold_col + 1, 0.50)
lists.cell(2, threshold_col + 1).number_format = "0%"
lists.cell(3, threshold_col + 1).number_format = "0%"
for cell in lists[1]:
    cell.font = Font(bold=True, color=WHITE)
    cell.fill = PatternFill("solid", fgColor=NAVY)
define_name(wb, "ReadyThreshold", "Lists", get_column_letter(threshold_col + 1), 2, 2)
define_name(wb, "PartialThreshold", "Lists", get_column_letter(threshold_col + 1), 3, 3)

# -----------------------------------------------------------------------------
# README
# -----------------------------------------------------------------------------
set_title(readme, "WS1 Campaign Discovery & Gap Analysis", "Pilot workbook for mapping the current WS1 campaign landscape before requesting performance data.", 8)
readme.merge_cells("A4:H5")
readme["A4"] = "This workbook is for campaign discovery and gap analysis only. It is not yet the final Promotion Performance Report."
readme["A4"].font = Font(name="Aptos", size=13, bold=True, color=NAVY)
readme["A4"].fill = PatternFill("solid", fgColor=CREAM)
readme["A4"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
section(readme, 7, "RECOMMENDED WORKFLOW", 1, 8)
workflow = [
    "Collect all active and recently completed WS1 campaigns.",
    "Confirm the campaign source and owner.",
    "Confirm the original campaign objective.",
    "Confirm the intended target segment.",
    "Document the actual segment-selection criteria.",
    "Record the campaign mechanic, timing and channel.",
    "Record what result data is available.",
    "Identify gaps.",
    "Assign clarification actions.",
    "Select suitable campaigns for future performance analysis.",
]
for i, text_value in enumerate(workflow, start=8):
    readme.cell(i, 1, i - 7)
    readme.cell(i, 2, text_value)
    readme.merge_cells(start_row=i, start_column=2, end_row=i, end_column=8)
    readme.cell(i, 1).font = Font(bold=True, color=GOLD)
    readme.cell(i, 2).alignment = Alignment(wrap_text=True)
    readme.cell(i, 2).border = border_bottom

section(readme, 19, "POSSIBLE INFORMATION SOURCES", 1, 8)
sources = ["Stakeholders", "Sales", "VIP or VM teams", "CK Analytics or CRM", "Promotion", "Management"]
for idx, source in enumerate(sources, start=20):
    readme.cell(idx, 1, "•")
    readme.cell(idx, 2, source)
readme.merge_cells("D20:H25")
readme["D20"] = "Important starting rule: VM weekly promotions should initially be categorised as Retention until the original objective is confirmed."
readme["D20"].fill = PatternFill("solid", fgColor=AMBER)
readme["D20"].font = Font(bold=True, color=AMBER_TEXT)
readme["D20"].alignment = Alignment(wrap_text=True, vertical="center")

section(readme, 27, "KEY DEFINITIONS", 1, 8)
definitions = [
    ("Campaign objective", "The specific business outcome the campaign is intended to support."),
    ("Target segment", "The player group the campaign is intended to reach."),
    ("Segment criteria", "The actual rules used to select eligible players."),
    ("Acquisition", "Campaigns intended to bring in new players or first depositors."),
    ("Retention", "Campaigns intended to keep existing active players engaged."),
    ("Reactivation", "Campaigns intended to bring inactive or lapsed players back."),
    ("Cross-product", "Campaigns intended to move players from one product to another."),
    ("VIP progression", "Campaigns intended to support movement to a higher value or VIP tier."),
    ("Campaign source", "The team or stakeholder that initiated the campaign."),
    ("Success measure", "The metric that will show whether the campaign achieved its objective."),
    ("Data availability", "Whether the information needed for analysis exists and can be accessed."),
    ("Objective clarity", "How clearly the campaign purpose is documented."),
    ("Segment clarity", "How clearly the target group and selection rules are documented."),
    ("Measurement readiness", "Whether sufficient result data and ownership exist for analysis."),
    ("Campaign overlap", "Two or more campaigns targeting similar players with similar mechanics."),
    ("Gap severity", "The urgency and likely business impact of missing or unclear information."),
]
for row, (term, definition) in enumerate(definitions, start=28):
    readme.cell(row, 1, term)
    readme.merge_cells(start_row=row, start_column=2, end_row=row, end_column=8)
    readme.cell(row, 2, definition)
    readme.cell(row, 1).fill = PatternFill("solid", fgColor=CREAM)
    readme.cell(row, 1).font = Font(bold=True, color=TEXT)
    readme.cell(row, 2).fill = PatternFill("solid", fgColor=WHITE)
    readme.cell(row, 2).alignment = Alignment(wrap_text=True)
    readme.cell(row, 1).border = border_bottom
    readme.cell(row, 2).border = border_bottom

section(readme, 45, "COLOUR KEY", 1, 8)
legend = [("Manual input", INPUT), ("Calculated field", CALC), ("Missing / urgent", RED), ("External confirmation", EXTERNAL), ("Ready / complete", GREEN)]
for idx, (label, color) in enumerate(legend, start=46):
    readme.cell(idx, 1, label)
    readme.cell(idx, 2, " ")
    readme.cell(idx, 2).fill = PatternFill("solid", fgColor=color)
readme.column_dimensions["A"].width = 25
for col in "BCDEFGH":
    readme.column_dimensions[col].width = 17
readme.freeze_panes = "A7"

# -----------------------------------------------------------------------------
# Campaign Inventory
# -----------------------------------------------------------------------------
identification = ["Campaign ID", "Brand", "Campaign name", "Campaign source", "Requesting stakeholder or team", "Campaign owner", "Promo PIC", "Campaign status", "Campaign start date", "Campaign end date", "Frequency", "Last reviewed date"]
strategy = ["Campaign category", "Stated campaign objective", "Objective source", "Objective clarity", "Expected player behaviour", "Planned success measure", "Planned target", "Business reason for campaign"]
target = ["Stated target segment", "Actual segment criteria", "Player activity status", "Inactive-day criteria", "Deposit-level criteria", "VIP-tier criteria", "Product-behaviour criteria", "Source product", "Target product", "Geographic or currency segment", "Exclusion criteria", "Estimated targeted players", "Segment clarity"]
mechanic = ["Bonus type", "Bonus percentage", "Minimum deposit", "Maximum bonus cap", "Turnover requirement", "Maximum claims per player", "Campaign budget", "Currency", "Delivery channel", "Promo code", "Recurring or one-time", "Abuse control", "QC status"]
measurement = ["Result data available", "Claim data available", "Promo-cost data available", "Deposit data available", "NGR data available", "Reactivation data available", "Cross-product data available", "Post-campaign behaviour available", "Data source", "Data owner", "Data last updated", "Measurement readiness"]
calculated = ["Objective completeness score", "Segment completeness score", "Measurement readiness score", "Overall readiness"]
gaps = ["Objective gap", "Segment gap", "Mechanic gap", "Ownership gap", "Data gap", "Duplicate or overlap risk", "Main gap category", "Gap severity", "Recommended next action", "Clarification required from", "Clarification status", "Notes"]
inv_headers = identification + strategy + target + mechanic + measurement + calculated + gaps
inv_col = {h: i + 1 for i, h in enumerate(inv_headers)}
last_inv_col = len(inv_headers)
set_title(inventory, "WS1 Campaign Inventory", "One row per campaign. Yellow cells are manual inputs, blue cells are calculated, and orange fields usually require external confirmation.", last_inv_col)

# group labels
group_starts = []
pos = 1
for label, group in [("IDENTIFICATION", identification), ("CAMPAIGN STRATEGY", strategy), ("TARGET SEGMENT", target), ("CAMPAIGN MECHANIC", mechanic), ("MEASUREMENT", measurement), ("CALCULATED READINESS", calculated), ("GAP ANALYSIS", gaps)]:
    start = pos
    end = pos + len(group) - 1
    inventory.merge_cells(start_row=4, start_column=start, end_row=4, end_column=end)
    c = inventory.cell(4, start, label)
    c.fill = PatternFill("solid", fgColor=GOLD)
    c.font = Font(bold=True, color=WHITE)
    c.alignment = Alignment(horizontal="center")
    group_starts.append((start, end))
    pos = end + 1
for idx, h in enumerate(inv_headers, start=1):
    inventory.cell(6, idx, h)
style_header_row(inventory, 6, last_inv_col)

calculated_indices = [inv_col[h] for h in calculated] + [inv_col["Measurement readiness"]]
external_headers = ["Requesting stakeholder or team", "Campaign owner", "Stated campaign objective", "Objective source", "Stated target segment", "Actual segment criteria", "Result data available", "Data source", "Data owner", "Clarification required from"]
external_indices = [inv_col[h] for h in external_headers]
style_table_body(inventory, 7, MAX_DATA_ROW, 1, last_inv_col, calculated_indices, external_indices)

samples = [
    {
        "Campaign ID": "WS1-SMP-001", "Brand": "WS1", "Campaign name": "SAMPLE – Active Player Weekly Reload", "Campaign source": "Promotion", "Requesting stakeholder or team": "Promotion", "Campaign owner": "Promotion", "Promo PIC": "Promo PIC A", "Campaign status": "Live", "Campaign start date": date(2026, 7, 1), "Campaign end date": date(2026, 9, 30), "Frequency": "Weekly", "Last reviewed date": date(2026, 7, 15),
        "Campaign category": "Retention", "Stated campaign objective": "Maintain weekly deposit activity among active low-value players", "Objective source": "Campaign Brief", "Objective clarity": "Clear", "Expected player behaviour": "Make a qualifying weekly deposit", "Planned success measure": "Claim-to-deposit rate and repeat deposit behaviour", "Planned target": "Improve versus prior four-week baseline", "Business reason for campaign": "Support ongoing player activity",
        "Stated target segment": "Active low-value players", "Actual segment criteria": "Active in last 7 days and average deposit below RM100", "Player activity status": "Active", "Deposit-level criteria": "Below RM100 average deposit", "Product-behaviour criteria": "Deposited at least twice in prior 30 days", "Source product": "Multi-Product", "Target product": "Multi-Product", "Geographic or currency segment": "MYR", "Exclusion criteria": "Exclude bonus-abuse flags and VIP Gold+", "Estimated targeted players": 2500, "Segment clarity": "Clear",
        "Bonus type": "Reload Bonus", "Bonus percentage": 20, "Minimum deposit": 50, "Maximum bonus cap": 100, "Turnover requirement": "8x bonus", "Maximum claims per player": 1, "Campaign budget": 25000, "Currency": "MYR", "Delivery channel": "Multi-Channel", "Promo code": "SMPWEEK20", "Recurring or one-time": "Recurring", "Abuse control": "Exclude flagged accounts; one claim per week", "QC status": "Passed",
        "Result data available": "Available", "Claim data available": "Available", "Promo-cost data available": "Available", "Deposit data available": "Available", "NGR data available": "Available", "Reactivation data available": "Available", "Cross-product data available": "Partial", "Post-campaign behaviour available": "Available", "Data source": "CK Analytics CRM", "Data owner": "CK Analytics", "Data last updated": date(2026, 7, 20),
        "Objective gap": "None", "Segment gap": "None", "Mechanic gap": "None", "Ownership gap": "None", "Data gap": "None", "Duplicate or overlap risk": "Low", "Main gap category": "No material discovery gap", "Gap severity": "Low", "Recommended next action": "Prepare Performance Analysis", "Clarification required from": "CK Analytics", "Clarification status": "Received", "Notes": SAMPLE_LABEL,
    },
    {
        "Campaign ID": "WS1-SMP-002", "Brand": "WS1", "Campaign name": "SAMPLE – VM Weekly Reward", "Campaign source": "VIP VM", "Requesting stakeholder or team": "VIP", "Campaign owner": "VIP", "Promo PIC": "Promo PIC B", "Campaign status": "Live", "Campaign start date": date(2026, 6, 1), "Campaign end date": date(2026, 12, 31), "Frequency": "Weekly", "Last reviewed date": date(2026, 7, 10),
        "Campaign category": "Retention", "Stated campaign objective": "Retain existing VM-managed players", "Objective source": "Stakeholder Confirmation", "Objective clarity": "Partly Clear", "Expected player behaviour": "Remain active and deposit weekly", "Planned success measure": "Weekly active and depositing VM players", "Planned target": "Not documented", "Business reason for campaign": "Support VM relationship management",
        "Stated target segment": "VM weekly player list", "Actual segment criteria": "List supplied weekly by VM team", "Player activity status": "Active", "VIP-tier criteria": "Multiple", "Product-behaviour criteria": "Managed by VM", "Source product": "Multi-Product", "Target product": "Multi-Product", "Geographic or currency segment": "MYR", "Exclusion criteria": "Not documented", "Estimated targeted players": 300, "Segment clarity": "Partly Clear",
        "Bonus type": "Free Credit", "Minimum deposit": 0, "Maximum bonus cap": 100, "Turnover requirement": "1x", "Maximum claims per player": 1, "Campaign budget": 10000, "Currency": "MYR", "Delivery channel": "VIP Direct Contact", "Promo code": "SMPVMWEEK", "Recurring or one-time": "Recurring", "Abuse control": "VM approval", "QC status": "Passed with Conditions",
        "Result data available": "Partial", "Claim data available": "Available", "Promo-cost data available": "Available", "Deposit data available": "Partial", "NGR data available": "Pending", "Reactivation data available": "Not Available", "Cross-product data available": "Not Available", "Post-campaign behaviour available": "Pending", "Data source": "VIP VM", "Data owner": "VIP", "Data last updated": date(2026, 7, 18),
        "Objective gap": "Planned target missing", "Segment gap": "Weekly list criteria not documented", "Mechanic gap": "Review Required", "Ownership gap": "None", "Data gap": "Post-campaign behaviour pending", "Duplicate or overlap risk": "Medium", "Main gap category": "Success measure missing", "Gap severity": "Medium", "Recommended next action": "Confirm Objective", "Clarification required from": "VIP", "Clarification status": "Requested", "Notes": SAMPLE_LABEL,
    },
    {
        "Campaign ID": "WS1-SMP-003", "Brand": "WS1", "Campaign name": "SAMPLE – 15–30 Day Reactivation", "Campaign source": "Sales", "Requesting stakeholder or team": "Sales", "Campaign owner": "Sales", "Promo PIC": "Promo PIC A", "Campaign status": "Completed", "Campaign start date": date(2026, 6, 15), "Campaign end date": date(2026, 6, 21), "Frequency": "Monthly", "Last reviewed date": date(2026, 7, 5),
        "Campaign category": "Reactivation", "Stated campaign objective": "Reactivate players inactive for 15–30 days", "Objective source": "Campaign Brief", "Objective clarity": "Clear", "Expected player behaviour": "Return, claim and deposit within 72 hours", "Planned success measure": "Reactivated depositing players", "Planned target": "Improve against control group", "Business reason for campaign": "Recover recently lapsed depositors",
        "Stated target segment": "Inactive 15–30 days", "Actual segment criteria": "No login or deposit for 15–30 days; deposited in prior 90 days", "Player activity status": "Inactive", "Inactive-day criteria": "15–30 days", "Deposit-level criteria": "At least one deposit in prior 90 days", "Product-behaviour criteria": "Previously active depositor", "Source product": "Multi-Product", "Target product": "Multi-Product", "Geographic or currency segment": "MYR", "Exclusion criteria": "Exclude self-excluded, abuse and active bonus holders", "Estimated targeted players": 1800, "Segment clarity": "Clear",
        "Bonus type": "Free Credit", "Minimum deposit": 0, "Maximum bonus cap": 30, "Turnover requirement": "5x", "Maximum claims per player": 1, "Campaign budget": 30000, "Currency": "MYR", "Delivery channel": "SMS", "Promo code": "SMPREACT30", "Recurring or one-time": "Recurring", "Abuse control": "One claim; exclude flagged players", "QC status": "Passed",
        "Result data available": "Partial", "Claim data available": "Available", "Promo-cost data available": "Available", "Deposit data available": "Available", "NGR data available": "Pending", "Reactivation data available": "Available", "Cross-product data available": "Not Available", "Post-campaign behaviour available": "Partial", "Data source": "CK Analytics CRM", "Data owner": "CK Analytics", "Data last updated": date(2026, 7, 1),
        "Objective gap": "None", "Segment gap": "None", "Mechanic gap": "None", "Ownership gap": "None", "Data gap": "NGR pending; do not calculate ROI", "Duplicate or overlap risk": "Low", "Main gap category": "Result data missing", "Gap severity": "Medium", "Recommended next action": "Request Result Data", "Clarification required from": "CK Analytics", "Clarification status": "In Progress", "Notes": SAMPLE_LABEL,
    },
    {
        "Campaign ID": "WS1-SMP-004", "Brand": "WS1", "Campaign name": "SAMPLE – Welcome Back Players", "Campaign source": "Stakeholder", "Requesting stakeholder or team": "Stakeholder", "Campaign owner": "", "Promo PIC": "Promo PIC B", "Campaign status": "Scheduled", "Campaign start date": date(2026, 8, 1), "Campaign end date": date(2026, 8, 7), "Frequency": "Ad Hoc", "Last reviewed date": "",
        "Campaign category": "Reactivation", "Stated campaign objective": "Bring inactive players back", "Objective source": "Inferred", "Objective clarity": "Partly Clear", "Expected player behaviour": "Return and deposit", "Planned success measure": "", "Planned target": "", "Business reason for campaign": "Increase active players",
        "Stated target segment": "Inactive players", "Actual segment criteria": "Not documented", "Player activity status": "Inactive", "Inactive-day criteria": "", "Deposit-level criteria": "", "VIP-tier criteria": "", "Product-behaviour criteria": "", "Source product": "Unknown", "Target product": "Multi-Product", "Geographic or currency segment": "MYR", "Exclusion criteria": "", "Estimated targeted players": "", "Segment clarity": "Unclear",
        "Bonus type": "Deposit Bonus", "Bonus percentage": 50, "Minimum deposit": 50, "Maximum bonus cap": 300, "Turnover requirement": "10x", "Maximum claims per player": 1, "Campaign budget": 50000, "Currency": "MYR", "Delivery channel": "Multi-Channel", "Promo code": "SMPBACK50", "Recurring or one-time": "One-Time", "Abuse control": "Unknown", "QC status": "Not Checked",
        "Result data available": "Pending", "Claim data available": "Pending", "Promo-cost data available": "Pending", "Deposit data available": "Pending", "NGR data available": "Pending", "Reactivation data available": "Pending", "Cross-product data available": "Unknown", "Post-campaign behaviour available": "Pending", "Data source": "Unknown", "Data owner": "",
        "Objective gap": "Success measure and target missing", "Segment gap": "Inactive-day criteria missing", "Mechanic gap": "Review Required", "Ownership gap": "Campaign owner missing", "Data gap": "Data owner and result plan missing", "Duplicate or overlap risk": "Medium", "Main gap category": "Segment criteria missing", "Gap severity": "Critical", "Recommended next action": "Define Segment Criteria", "Clarification required from": "Stakeholder", "Clarification status": "Not Started", "Notes": SAMPLE_LABEL,
    },
    {
        "Campaign ID": "WS1-SMP-005", "Brand": "WS1", "Campaign name": "SAMPLE – Sports to Casino Bridge", "Campaign source": "Sales", "Requesting stakeholder or team": "Sales", "Campaign owner": "Sales", "Promo PIC": "Promo PIC A", "Campaign status": "Completed", "Campaign start date": date(2026, 6, 20), "Campaign end date": date(2026, 6, 30), "Frequency": "One-Time", "Last reviewed date": date(2026, 7, 8),
        "Campaign category": "Cross-Product", "Stated campaign objective": "Encourage active Sports-only players to make a first Casino deposit", "Objective source": "Campaign Brief", "Objective clarity": "Clear", "Expected player behaviour": "Open Casino and complete a qualifying Casino deposit", "Planned success measure": "Sports-only to Casino depositing conversion", "Planned target": "Test against a holdout group", "Business reason for campaign": "Increase multi-product engagement",
        "Stated target segment": "Sports-only active depositors", "Actual segment criteria": "Sports deposit in prior 30 days and no Casino wager in prior 90 days", "Player activity status": "Active", "Deposit-level criteria": "Sports deposit ≥ RM100 in prior 30 days", "Product-behaviour criteria": "Sports-only; no Casino activity 90 days", "Source product": "Sports", "Target product": "Casino", "Geographic or currency segment": "MYR", "Exclusion criteria": "Exclude Casino-active and bonus-abuse players", "Estimated targeted players": 1200, "Segment clarity": "Clear",
        "Bonus type": "Deposit Bonus", "Bonus percentage": 30, "Minimum deposit": 50, "Maximum bonus cap": 150, "Turnover requirement": "12x", "Maximum claims per player": 1, "Campaign budget": 35000, "Currency": "MYR", "Delivery channel": "MT", "Promo code": "SMPS2C30", "Recurring or one-time": "One-Time", "Abuse control": "First Casino deposit only", "QC status": "Passed",
        "Result data available": "Partial", "Claim data available": "Available", "Promo-cost data available": "Available", "Deposit data available": "Available", "NGR data available": "Pending", "Reactivation data available": "Not Available", "Cross-product data available": "Available", "Post-campaign behaviour available": "Partial", "Data source": "CK Analytics CRM", "Data owner": "CK Analytics", "Data last updated": date(2026, 7, 5),
        "Objective gap": "None", "Segment gap": "None", "Mechanic gap": "None", "Ownership gap": "None", "Data gap": "NGR and longer-term behaviour pending", "Duplicate or overlap risk": "Low", "Main gap category": "Result data missing", "Gap severity": "Medium", "Recommended next action": "Request Result Data", "Clarification required from": "CK Analytics", "Clarification status": "In Progress", "Notes": SAMPLE_LABEL,
    },
    {
        "Campaign ID": "WS1-SMP-006", "Brand": "WS1", "Campaign name": "SAMPLE – Gold to Platinum Milestone", "Campaign source": "VIP VM", "Requesting stakeholder or team": "VIP", "Campaign owner": "VIP", "Promo PIC": "Promo PIC C", "Campaign status": "Draft", "Campaign start date": date(2026, 8, 15), "Campaign end date": date(2026, 8, 31), "Frequency": "Monthly", "Last reviewed date": date(2026, 7, 19),
        "Campaign category": "VIP Progression", "Stated campaign objective": "Encourage eligible Gold players to reach Platinum activity milestones", "Objective source": "Stakeholder Confirmation", "Objective clarity": "Clear", "Expected player behaviour": "Increase qualified deposit and activity during the milestone period", "Planned success measure": "Eligible players progressing to Platinum", "Planned target": "Confirm with VIP", "Business reason for campaign": "Support valuable-player progression",
        "Stated target segment": "Gold-tier players near Platinum threshold", "Actual segment criteria": "Gold tier; within 20% of Platinum monthly threshold", "Player activity status": "Active", "Deposit-level criteria": "Within 20% of target", "VIP-tier criteria": "Gold", "Product-behaviour criteria": "Active in prior 14 days", "Source product": "Multi-Product", "Target product": "Multi-Product", "Geographic or currency segment": "MYR", "Exclusion criteria": "Exclude downgraded or abuse-flagged accounts", "Estimated targeted players": 180, "Segment clarity": "Clear",
        "Bonus type": "Mystery Reward", "Minimum deposit": 500, "Maximum bonus cap": 500, "Turnover requirement": "8x", "Maximum claims per player": 1, "Campaign budget": 30000, "Currency": "MYR", "Delivery channel": "VIP Direct Contact", "Promo code": "SMPG2P", "Recurring or one-time": "One-Time", "Abuse control": "VIP approval and tier validation", "QC status": "Not Checked",
        "Result data available": "Pending", "Claim data available": "Pending", "Promo-cost data available": "Pending", "Deposit data available": "Pending", "NGR data available": "Pending", "Reactivation data available": "Not Available", "Cross-product data available": "Not Available", "Post-campaign behaviour available": "Pending", "Data source": "VIP", "Data owner": "VIP",
        "Objective gap": "Planned target requires confirmation", "Segment gap": "None", "Mechanic gap": "Review Required", "Ownership gap": "None", "Data gap": "Result-data extract not agreed", "Duplicate or overlap risk": "Low", "Main gap category": "Success measure missing", "Gap severity": "High", "Recommended next action": "Confirm Objective", "Clarification required from": "VIP", "Clarification status": "Requested", "Notes": SAMPLE_LABEL,
    },
    {
        "Campaign ID": "WS1-SMP-007", "Brand": "WS1", "Campaign name": "SAMPLE – Payday Mega Bonus", "Campaign source": "Management", "Requesting stakeholder or team": "Management", "Campaign owner": "Sales", "Promo PIC": "Promo PIC B", "Campaign status": "Scheduled", "Campaign start date": date(2026, 7, 25), "Campaign end date": date(2026, 7, 31), "Frequency": "Payday", "Last reviewed date": date(2026, 7, 18),
        "Campaign category": "Payday", "Stated campaign objective": "Increase payday deposits", "Objective source": "Campaign Brief", "Objective clarity": "Partly Clear", "Expected player behaviour": "Deposit during payday week", "Planned success measure": "Total deposits and claims", "Planned target": "Not segmented", "Business reason for campaign": "Capture payday activity",
        "Stated target segment": "All players", "Actual segment criteria": "All contactable players", "Player activity status": "Unknown", "Source product": "Multi-Product", "Target product": "Multi-Product", "Geographic or currency segment": "MYR", "Exclusion criteria": "Self-excluded only", "Estimated targeted players": 50000, "Segment clarity": "Too Broad",
        "Bonus type": "Deposit Bonus", "Bonus percentage": 100, "Minimum deposit": 100, "Maximum bonus cap": 1000, "Turnover requirement": "15x", "Maximum claims per player": 1, "Campaign budget": 250000, "Currency": "MYR", "Delivery channel": "Multi-Channel", "Promo code": "SMPPAY100", "Recurring or one-time": "Recurring", "Abuse control": "Basic eligibility only", "QC status": "Recheck Required",
        "Result data available": "Pending", "Claim data available": "Pending", "Promo-cost data available": "Pending", "Deposit data available": "Pending", "NGR data available": "Pending", "Reactivation data available": "Unknown", "Cross-product data available": "Unknown", "Post-campaign behaviour available": "Pending", "Data source": "CK Analytics CRM", "Data owner": "CK Analytics",
        "Objective gap": "Objective too general", "Segment gap": "All players is too broad", "Mechanic gap": "High cap requires review", "Ownership gap": "None", "Data gap": "Result plan pending", "Duplicate or overlap risk": "High", "Main gap category": "Segment too broad", "Gap severity": "Critical", "Recommended next action": "Run Controlled Test", "Clarification required from": "Management", "Clarification status": "In Progress", "Notes": SAMPLE_LABEL,
    },
    {
        "Campaign ID": "WS1-SMP-008", "Brand": "WS1", "Campaign name": "SAMPLE – Standard Recurring Bonus", "Campaign source": "Promotion", "Requesting stakeholder or team": "Promotion", "Campaign owner": "", "Promo PIC": "Promo PIC A", "Campaign status": "Live", "Campaign start date": date(2025, 1, 1), "Campaign end date": "", "Frequency": "Always-On", "Last reviewed date": "",
        "Campaign category": "Standard Recurring Bonus", "Stated campaign objective": "", "Objective source": "Not Documented", "Objective clarity": "Not Provided", "Expected player behaviour": "Deposit and claim", "Planned success measure": "", "Planned target": "", "Business reason for campaign": "Historical recurring offer",
        "Stated target segment": "Eligible players", "Actual segment criteria": "Platform default eligibility", "Player activity status": "Unknown", "Source product": "Casino", "Target product": "Casino", "Geographic or currency segment": "MYR", "Exclusion criteria": "Not documented", "Estimated targeted players": "", "Segment clarity": "Too Broad",
        "Bonus type": "Reload Bonus", "Bonus percentage": 20, "Minimum deposit": 30, "Maximum bonus cap": 200, "Turnover requirement": "10x", "Maximum claims per player": 1, "Campaign budget": "", "Currency": "MYR", "Delivery channel": "Website Dialog", "Promo code": "SMPSTD20", "Recurring or one-time": "Recurring", "Abuse control": "System eligibility", "QC status": "Unknown",
        "Result data available": "Unknown", "Claim data available": "Unknown", "Promo-cost data available": "Unknown", "Deposit data available": "Unknown", "NGR data available": "Unknown", "Reactivation data available": "Unknown", "Cross-product data available": "Unknown", "Post-campaign behaviour available": "Unknown", "Data source": "Unknown", "Data owner": "",
        "Objective gap": "Objective not documented", "Segment gap": "Target is too broad", "Mechanic gap": "Review Required", "Ownership gap": "Campaign owner missing", "Data gap": "No documented result data", "Duplicate or overlap risk": "High", "Main gap category": "Campaign purpose no longer clear", "Gap severity": "Critical", "Recommended next action": "Consider Sunset", "Clarification required from": "Management", "Clarification status": "Not Started", "Notes": SAMPLE_LABEL,
    },
    {
        "Campaign ID": "WS1-SMP-009", "Brand": "WS1", "Campaign name": "SAMPLE – Weekend Reload A", "Campaign source": "Sales", "Requesting stakeholder or team": "Sales", "Campaign owner": "Sales", "Promo PIC": "Promo PIC B", "Campaign status": "Live", "Campaign start date": date(2026, 7, 1), "Campaign end date": date(2026, 8, 31), "Frequency": "Weekly", "Last reviewed date": date(2026, 7, 12),
        "Campaign category": "Retention", "Stated campaign objective": "Increase weekend deposits from active Casino players", "Objective source": "Campaign Brief", "Objective clarity": "Clear", "Expected player behaviour": "Deposit on Friday to Sunday", "Planned success measure": "Weekend depositing players", "Planned target": "Increase versus weekday baseline", "Business reason for campaign": "Strengthen weekend activity",
        "Stated target segment": "Active Casino depositors", "Actual segment criteria": "Casino deposit in prior 30 days", "Player activity status": "Active", "Deposit-level criteria": "Any deposit prior 30 days", "Product-behaviour criteria": "Casino active", "Source product": "Casino", "Target product": "Casino", "Geographic or currency segment": "MYR", "Exclusion criteria": "Exclude active bonus holders", "Estimated targeted players": 7000, "Segment clarity": "Clear",
        "Bonus type": "Reload Bonus", "Bonus percentage": 30, "Minimum deposit": 50, "Maximum bonus cap": 200, "Turnover requirement": "10x", "Maximum claims per player": 1, "Campaign budget": 75000, "Currency": "MYR", "Delivery channel": "SMS", "Promo code": "SMPWKND30A", "Recurring or one-time": "Recurring", "Abuse control": "One claim per weekend", "QC status": "Passed",
        "Result data available": "Partial", "Claim data available": "Available", "Promo-cost data available": "Available", "Deposit data available": "Available", "NGR data available": "Pending", "Reactivation data available": "Not Available", "Cross-product data available": "Not Available", "Post-campaign behaviour available": "Partial", "Data source": "CK Analytics CRM", "Data owner": "CK Analytics", "Data last updated": date(2026, 7, 18),
        "Objective gap": "None", "Segment gap": "None", "Mechanic gap": "Overlap review required", "Ownership gap": "None", "Data gap": "NGR pending", "Duplicate or overlap risk": "High", "Main gap category": "Duplicate or overlapping mechanic", "Gap severity": "High", "Recommended next action": "Merge Similar Campaigns", "Clarification required from": "Sales", "Clarification status": "Requested", "Notes": SAMPLE_LABEL,
    },
    {
        "Campaign ID": "WS1-SMP-010", "Brand": "WS1", "Campaign name": "SAMPLE – Weekend Reload B", "Campaign source": "Stakeholder", "Requesting stakeholder or team": "Stakeholder", "Campaign owner": "Stakeholder", "Promo PIC": "Promo PIC C", "Campaign status": "Scheduled", "Campaign start date": date(2026, 7, 25), "Campaign end date": date(2026, 8, 31), "Frequency": "Weekly", "Last reviewed date": date(2026, 7, 16),
        "Campaign category": "General Promotion", "Stated campaign objective": "Drive weekend deposits", "Objective source": "Campaign Brief", "Objective clarity": "Partly Clear", "Expected player behaviour": "Deposit during weekend", "Planned success measure": "Claims", "Planned target": "Not documented", "Business reason for campaign": "Weekend activity",
        "Stated target segment": "Active players", "Actual segment criteria": "Active in prior 30 days", "Player activity status": "Active", "Deposit-level criteria": "Any", "Product-behaviour criteria": "Casino active", "Source product": "Casino", "Target product": "Casino", "Geographic or currency segment": "MYR", "Exclusion criteria": "Exclude active bonus holders", "Estimated targeted players": 9000, "Segment clarity": "Partly Clear",
        "Bonus type": "Reload Bonus", "Bonus percentage": 25, "Minimum deposit": 50, "Maximum bonus cap": 250, "Turnover requirement": "10x", "Maximum claims per player": 1, "Campaign budget": 80000, "Currency": "MYR", "Delivery channel": "MT", "Promo code": "SMPWKND25B", "Recurring or one-time": "Recurring", "Abuse control": "One claim per weekend", "QC status": "Passed with Conditions",
        "Result data available": "Pending", "Claim data available": "Pending", "Promo-cost data available": "Pending", "Deposit data available": "Pending", "NGR data available": "Pending", "Reactivation data available": "Not Available", "Cross-product data available": "Not Available", "Post-campaign behaviour available": "Pending", "Data source": "Unknown", "Data owner": "",
        "Objective gap": "Planned target missing", "Segment gap": "Criteria overlap with WS1-SMP-009", "Mechanic gap": "Overlap review required", "Ownership gap": "Data owner missing", "Data gap": "Result data not agreed", "Duplicate or overlap risk": "High", "Main gap category": "Duplicate or overlapping mechanic", "Gap severity": "High", "Recommended next action": "Merge Similar Campaigns", "Clarification required from": "Stakeholder", "Clarification status": "Not Started", "Notes": SAMPLE_LABEL,
    },
    {
        "Campaign ID": "WS1-SMP-011", "Brand": "WS1", "Campaign name": "SAMPLE – New Depositor Follow-Up", "Campaign source": "CK Analytics CRM", "Requesting stakeholder or team": "CK Analytics", "Campaign owner": "Promotion", "Promo PIC": "Promo PIC A", "Campaign status": "Completed", "Campaign start date": date(2026, 6, 1), "Campaign end date": date(2026, 6, 30), "Frequency": "Monthly", "Last reviewed date": date(2026, 7, 10),
        "Campaign category": "Retention", "Stated campaign objective": "Encourage new depositors to make a second deposit within 14 days", "Objective source": "Campaign Brief", "Objective clarity": "Clear", "Expected player behaviour": "Make a repeat deposit", "Planned success measure": "Second-deposit rate within 14 days", "Planned target": "Improve versus historical baseline", "Business reason for campaign": "Improve early retention",
        "Stated target segment": "New depositors", "Actual segment criteria": "First deposit in prior 7 days and no second deposit", "Player activity status": "New", "Inactive-day criteria": "0–7 days since first deposit", "Deposit-level criteria": "One completed deposit", "Product-behaviour criteria": "No repeat deposit", "Source product": "Multi-Product", "Target product": "Multi-Product", "Geographic or currency segment": "MYR", "Exclusion criteria": "Exclude self-excluded, bonus-abuse and duplicate accounts", "Estimated targeted players": 900, "Segment clarity": "Clear",
        "Bonus type": "Voucher", "Minimum deposit": 50, "Maximum bonus cap": 50, "Turnover requirement": "6x", "Maximum claims per player": 1, "Campaign budget": 15000, "Currency": "MYR", "Delivery channel": "Email", "Promo code": "SMP2NDDEP", "Recurring or one-time": "Recurring", "Abuse control": "One voucher after first deposit", "QC status": "Passed",
        "Result data available": "Not Available", "Claim data available": "Not Available", "Promo-cost data available": "Not Available", "Deposit data available": "Not Available", "NGR data available": "Not Available", "Reactivation data available": "Not Available", "Cross-product data available": "Not Available", "Post-campaign behaviour available": "Not Available", "Data source": "CK Analytics CRM", "Data owner": "CK Analytics", "Data last updated": "",
        "Objective gap": "None", "Segment gap": "None", "Mechanic gap": "None", "Ownership gap": "None", "Data gap": "Analytics result extract not received", "Duplicate or overlap risk": "Low", "Main gap category": "Result data missing", "Gap severity": "High", "Recommended next action": "Request Result Data", "Clarification required from": "CK Analytics", "Clarification status": "Requested", "Notes": SAMPLE_LABEL,
    },
    {
        "Campaign ID": "WS1-SMP-012", "Brand": "WS1", "Campaign name": "SAMPLE – Seasonal Acquisition Welcome", "Campaign source": "Management", "Requesting stakeholder or team": "Management", "Campaign owner": "Sales", "Promo PIC": "Promo PIC C", "Campaign status": "Draft", "Campaign start date": date(2026, 9, 1), "Campaign end date": date(2026, 9, 15), "Frequency": "Seasonal", "Last reviewed date": date(2026, 7, 20),
        "Campaign category": "Acquisition", "Stated campaign objective": "Support first deposits from newly registered players", "Objective source": "Campaign Brief", "Objective clarity": "Clear", "Expected player behaviour": "Complete first deposit", "Planned success measure": "First-depositor conversion", "Planned target": "To be confirmed with Sales", "Business reason for campaign": "Support seasonal new-player activity",
        "Stated target segment": "Registered non-depositors", "Actual segment criteria": "Registered within 14 days and no completed deposit", "Player activity status": "New", "Deposit-level criteria": "No deposit", "Product-behaviour criteria": "Registered; no funding", "Source product": "Unknown", "Target product": "Multi-Product", "Geographic or currency segment": "MYR", "Exclusion criteria": "Exclude duplicate and self-excluded accounts", "Estimated targeted players": 2000, "Segment clarity": "Clear",
        "Bonus type": "Deposit Bonus", "Bonus percentage": 100, "Minimum deposit": 50, "Maximum bonus cap": 200, "Turnover requirement": "15x", "Maximum claims per player": 1, "Campaign budget": 80000, "Currency": "MYR", "Delivery channel": "Multi-Channel", "Promo code": "SMPWELCOME", "Recurring or one-time": "One-Time", "Abuse control": "First deposit only; device and identity checks", "QC status": "Not Checked",
        "Result data available": "Pending", "Claim data available": "Pending", "Promo-cost data available": "Pending", "Deposit data available": "Pending", "NGR data available": "Pending", "Reactivation data available": "Not Available", "Cross-product data available": "Not Available", "Post-campaign behaviour available": "Pending", "Data source": "CK Analytics CRM", "Data owner": "CK Analytics",
        "Objective gap": "Planned target requires confirmation", "Segment gap": "None", "Mechanic gap": "Review Required", "Ownership gap": "None", "Data gap": "Result plan pending", "Duplicate or overlap risk": "Low", "Main gap category": "Success measure missing", "Gap severity": "Medium", "Recommended next action": "Confirm Objective", "Clarification required from": "Sales", "Clarification status": "Not Started", "Notes": SAMPLE_LABEL,
    },
]

for row_idx, sample in enumerate(samples, start=7):
    for header, value in sample.items():
        inventory.cell(row_idx, inv_col[header], value)

# Formulas: populate sample rows and future input rows.
def letter(header):
    return get_column_letter(inv_col[header])

for r in range(7, MAX_DATA_ROW + 1):
    cid = f"{letter('Campaign ID')}{r}"
    obj_fields = [letter("Stated campaign objective"), letter("Expected player behaviour"), letter("Planned success measure"), letter("Planned target")]
    inventory.cell(r, inv_col["Objective completeness score"], f'=IF({cid}="","",COUNTA({obj_fields[0]}{r},{obj_fields[1]}{r},{obj_fields[2]}{r},{obj_fields[3]}{r})/4)')
    behavioral = f'IF(COUNTA({letter("Player activity status")}{r},{letter("Inactive-day criteria")}{r},{letter("Deposit-level criteria")}{r},{letter("VIP-tier criteria")}{r},{letter("Product-behaviour criteria")}{r})>0,1,0)'
    inventory.cell(r, inv_col["Segment completeness score"], f'=IF({cid}="","",(COUNTA({letter("Stated target segment")}{r},{letter("Actual segment criteria")}{r},{letter("Exclusion criteria")}{r})+{behavioral})/4)')
    availability_range = f'{letter("Claim data available")}{r}:{letter("Reactivation data available")}{r}'
    data_owner = f'{letter("Data owner")}{r}'
    inventory.cell(r, inv_col["Measurement readiness score"], f'=IF({cid}="","",(COUNTIF({availability_range},"Available")+0.5*COUNTIF({availability_range},"Partial")+IF({data_owner}<>"",1,0))/6)')
    oscore = f'{letter("Objective completeness score")}{r}'
    sscore = f'{letter("Segment completeness score")}{r}'
    mscore = f'{letter("Measurement readiness score")}{r}'
    owner = f'{letter("Campaign owner")}{r}'
    inventory.cell(r, inv_col["Overall readiness"], f'=IF({cid}="","",IF(AND({owner}="",{data_owner}=""),"Blocked",IF(AND({oscore}>=ReadyThreshold,{sscore}>=ReadyThreshold,{mscore}>=ReadyThreshold),"Ready for Analysis",IF(AVERAGE({oscore},{sscore},{mscore})>=PartialThreshold,"Partially Ready","Discovery Required"))))')
    inventory.cell(r, inv_col["Measurement readiness"], f'=IF({cid}="","",IF({mscore}>=ReadyThreshold,"Ready",IF({mscore}>=PartialThreshold,"Partial","Not Ready")))')

for h in ["Campaign start date", "Campaign end date", "Last reviewed date", "Data last updated"]:
    for r in range(7, MAX_DATA_ROW + 1): inventory.cell(r, inv_col[h]).number_format = "yyyy-mm-dd"
for h in ["Bonus percentage"]:
    for r in range(7, MAX_DATA_ROW + 1): inventory.cell(r, inv_col[h]).number_format = '0"%"'
for h in ["Minimum deposit", "Maximum bonus cap", "Campaign budget"]:
    for r in range(7, MAX_DATA_ROW + 1): inventory.cell(r, inv_col[h]).number_format = '#,##0.00'
for h in calculated[:3]:
    for r in range(7, MAX_DATA_ROW + 1): inventory.cell(r, inv_col[h]).number_format = "0%"

# Header notes for complex fields
comments = {
    "Objective completeness score": "Calculated from objective, expected behaviour, success measure and planned target.",
    "Segment completeness score": "Calculated from target segment, actual criteria, at least one player/behaviour criterion, and exclusions.",
    "Measurement readiness score": "Available=1, Partial=0.5 for claim, promo cost, deposit, NGR and reactivation data, plus a named data owner.",
    "Overall readiness": "Uses editable thresholds stored on the hidden Lists tab; blocked when both campaign owner and data owner are missing.",
    "Duplicate or overlap risk": "Flag when campaigns target similar players with similar mechanics or timing.",
}
for h, comment in comments.items(): inventory.cell(6, inv_col[h]).comment = Comment(comment, "User")

# Data validations
validation_map = {
    "Brand": "Brands", "Campaign source": "CampaignSources", "Campaign status": "CampaignStatuses", "Frequency": "Frequencies",
    "Campaign category": "CampaignCategories", "Objective source": "ObjectiveSources", "Objective clarity": "ObjectiveClarityStatuses",
    "Player activity status": "PlayerActivityStatuses", "VIP-tier criteria": "VIPTiers", "Source product": "Products", "Target product": "Products", "Segment clarity": "SegmentClarityStatuses",
    "Bonus type": "BonusTypes", "Currency": "Currencies", "Delivery channel": "DeliveryChannels", "Recurring or one-time": "RecurringStatuses", "QC status": "QCStatuses",
    "Result data available": "DataAvailabilityStatuses", "Claim data available": "DataAvailabilityStatuses", "Promo-cost data available": "DataAvailabilityStatuses", "Deposit data available": "DataAvailabilityStatuses", "NGR data available": "DataAvailabilityStatuses", "Reactivation data available": "DataAvailabilityStatuses", "Cross-product data available": "DataAvailabilityStatuses", "Post-campaign behaviour available": "DataAvailabilityStatuses",
    "Mechanic gap": "MechanicGapStatuses", "Duplicate or overlap risk": "OverlapRisks", "Gap severity": "GapSeverities", "Recommended next action": "RecommendedActions", "Clarification status": "ClarificationStatuses",
}
for h, named in validation_map.items():
    col_letter = letter(h)
    add_list_validation(inventory, f"{col_letter}7:{col_letter}{MAX_DATA_ROW}", named)

add_table(inventory, f"A6:{get_column_letter(last_inv_col)}{MAX_DATA_ROW}", "tblCampaignInventory")
inventory.freeze_panes = "H7"
inventory.auto_filter.ref = f"A6:{get_column_letter(last_inv_col)}{MAX_DATA_ROW}"
inventory.sheet_view.zoomScale = 70

# Column widths
for idx, h in enumerate(inv_headers, start=1):
    width = 16
    if h in ["Campaign name", "Stated campaign objective", "Expected player behaviour", "Business reason for campaign", "Stated target segment", "Actual segment criteria", "Product-behaviour criteria", "Exclusion criteria", "Abuse control", "Objective gap", "Segment gap", "Data gap", "Recommended next action", "Notes"]:
        width = 24
    if h in ["Campaign ID", "Promo code", "Currency", "Brand"]: width = 14
    inventory.column_dimensions[get_column_letter(idx)].width = width

# Conditional formatting
data_range = f"A7:{get_column_letter(last_inv_col)}{MAX_DATA_ROW}"
inventory.conditional_formatting.add(f"{letter('Objective clarity')}7:{letter('Objective clarity')}{MAX_DATA_ROW}", FormulaRule(formula=[f'OR({letter("Objective clarity")}7="Unclear",{letter("Objective clarity")}7="Not Provided")'], fill=PatternFill("solid", fgColor=RED), font=Font(color=RED_TEXT, bold=True)))
inventory.conditional_formatting.add(f"{letter('Segment clarity')}7:{letter('Segment clarity')}{MAX_DATA_ROW}", FormulaRule(formula=[f'OR({letter("Segment clarity")}7="Too Broad",{letter("Segment clarity")}7="Unclear",{letter("Segment clarity")}7="Not Provided")'], fill=PatternFill("solid", fgColor=AMBER), font=Font(color=AMBER_TEXT, bold=True)))
inventory.conditional_formatting.add(f"{letter('Result data available')}7:{letter('Result data available')}{MAX_DATA_ROW}", FormulaRule(formula=[f'OR({letter("Result data available")}7="Pending",{letter("Result data available")}7="Not Available",{letter("Result data available")}7="Unknown")'], fill=PatternFill("solid", fgColor=RED), font=Font(color=RED_TEXT, bold=True)))
for h in ["Campaign owner", "Data owner"]:
    col_l = letter(h)
    inventory.conditional_formatting.add(f"{col_l}7:{col_l}{MAX_DATA_ROW}", FormulaRule(formula=[f'AND($A7<>"",{col_l}7="")'], fill=PatternFill("solid", fgColor=RED)))
inventory.conditional_formatting.add(f"{letter('Gap severity')}7:{letter('Gap severity')}{MAX_DATA_ROW}", FormulaRule(formula=[f'OR({letter("Gap severity")}7="High",{letter("Gap severity")}7="Critical")'], fill=PatternFill("solid", fgColor=RED), font=Font(color=RED_TEXT, bold=True)))
inventory.conditional_formatting.add(f"{letter('Overall readiness')}7:{letter('Overall readiness')}{MAX_DATA_ROW}", FormulaRule(formula=[f'{letter("Overall readiness")}7="Ready for Analysis"'], fill=PatternFill("solid", fgColor=GREEN), font=Font(color=GREEN_TEXT, bold=True)))

# -----------------------------------------------------------------------------
# Gap Summary
# -----------------------------------------------------------------------------
set_title(gap, "WS1 Campaign Gap Summary", "Management view of discovery gaps only. Figures are driven by the Campaign Inventory and synthetic samples must be replaced.", 18)
filters = [("A3", "Brand", "B3", "FilterAllBrands", "WS1"), ("C3", "Campaign category", "D3", "CampaignCategories", "All"), ("E3", "Campaign source", "F3", "CampaignSources", "All"), ("G3", "Campaign owner", "H3", "OwnerGroups", "All"), ("I3", "Gap severity", "J3", "GapSeverities", "All"), ("K3", "Campaign status", "L3", "CampaignStatuses", "All")]
# local All-enabled lists for filters
for name, base_name in [("FilterCategories", "CampaignCategories"), ("FilterSources", "CampaignSources"), ("FilterOwners", "OwnerGroups"), ("FilterSeverities", "GapSeverities"), ("FilterStatuses", "CampaignStatuses")]:
    # append in far columns on Lists with All first
    idx = lists.max_column + 1
    lists.cell(1, idx, name)
    lists.cell(1, idx).font = Font(bold=True, color=WHITE)
    lists.cell(1, idx).fill = PatternFill("solid", fgColor=NAVY)
    base_vals = list_values[base_name]
    values = ["All"] + base_vals
    for rr, val in enumerate(values, start=2): lists.cell(rr, idx, val)
    define_name(wb, name, "Lists", get_column_letter(idx), 2, len(values) + 1)
filter_named = ["FilterAllBrands", "FilterCategories", "FilterSources", "FilterOwners", "FilterSeverities", "FilterStatuses"]
for (label_cell, label, value_cell, _, default_value), named in zip(filters, filter_named):
    gap[label_cell] = label
    gap[label_cell].font = Font(bold=True, color=TEXT)
    gap[value_cell] = default_value
    gap[value_cell].fill = PatternFill("solid", fgColor=INPUT)
    add_list_validation(gap, value_cell, named)
for col in range(1, 13): gap.column_dimensions[get_column_letter(col)].width = 15

def inv_rng(h):
    l = letter(h)
    return f"'Campaign Inventory'!${l}$7:${l}${MAX_DATA_ROW}"

base_terms = [
    f'--((($B$3="All")+({inv_rng("Brand")}=$B$3))>0)',
    f'--((($D$3="All")+({inv_rng("Campaign category")}=$D$3))>0)',
    f'--((($F$3="All")+({inv_rng("Campaign source")}=$F$3))>0)',
    f'--((($H$3="All")+({inv_rng("Campaign owner")}=$H$3))>0)',
    f'--((($J$3="All")+({inv_rng("Gap severity")}=$J$3))>0)',
    f'--((($L$3="All")+({inv_rng("Campaign status")}=$L$3))>0)',
    f'--({inv_rng("Campaign ID")}<>"")',
]
def filtered_count(extra=None):
    terms = base_terms.copy()
    if extra: terms.extend(extra)
    return "=SUMPRODUCT(" + ",".join(terms) + ")"

kpis = [
    ("A5", "A6", "Total campaigns", filtered_count()),
    ("C5", "C6", "Active campaigns", filtered_count([f'--((({inv_rng("Campaign status")}="Live")+({inv_rng("Campaign status")}="Scheduled"))>0)'])),
    ("E5", "E6", "Clear objectives", filtered_count([f'--({inv_rng("Objective clarity")}="Clear")'])),
    ("G5", "G6", "Unclear objectives", filtered_count([f'--((({inv_rng("Objective clarity")}="Unclear")+({inv_rng("Objective clarity")}="Not Provided"))>0)'])),
    ("I5", "I6", "Clear segments", filtered_count([f'--({inv_rng("Segment clarity")}="Clear")'])),
    ("K5", "K6", "Broad / unclear segments", filtered_count([f'--((({inv_rng("Segment clarity")}="Too Broad")+({inv_rng("Segment clarity")}="Unclear")+({inv_rng("Segment clarity")}="Not Provided"))>0)'])),
    ("A8", "A9", "Complete result data", filtered_count([f'--({inv_rng("Result data available")}="Available")'])),
    ("C8", "C9", "Missing result data", filtered_count([f'--((({inv_rng("Result data available")}="Not Available")+({inv_rng("Result data available")}="Pending")+({inv_rng("Result data available")}="Unknown"))>0)'])),
    ("E8", "E9", "No clear owner", filtered_count([f'--({inv_rng("Campaign owner")}="")'])),
    ("G8", "G9", "Duplicate / overlap risk", filtered_count([f'--((({inv_rng("Duplicate or overlap risk")}="High")+({inv_rng("Duplicate or overlap risk")}="Medium"))>0)'])),
    ("I8", "I9", "Ready for analysis", filtered_count([f'--({inv_rng("Overall readiness")}="Ready for Analysis")'])),
    ("K8", "K9", "Critical gaps", filtered_count([f'--({inv_rng("Gap severity")}="Critical")'])),
]
for label_cell, value_cell, label_text, formula in kpis:
    card(gap, label_cell, value_cell, label_text, formula)
    gap.row_dimensions[5 if label_cell.endswith("5") else 8].height = 28
    gap.row_dimensions[6 if value_cell.endswith("6") else 9].height = 38

section(gap, 11, "GAP SUMMARY", 1, 6)
gap_headers = ["Gap category", "Number of campaigns", "Percentage of campaigns", "Example campaign", "Business impact", "Recommended response"]
for i, h in enumerate(gap_headers, start=1): gap.cell(12, i, h)
style_header_row(gap, 12, 6)
gap_categories = [
    ("Objective unclear", f'--((({inv_rng("Objective clarity")}="Unclear")+({inv_rng("Objective clarity")}="Not Provided"))>0)', "Campaign purpose cannot be evaluated", "Confirm the original objective"),
    ("Objective too general", f'--({inv_rng("Objective gap")}="Objective too general")', "Campaign may optimise the wrong behaviour", "Rewrite as a specific player outcome"),
    ("Segment too broad", f'--({inv_rng("Segment clarity")}="Too Broad")', "Cost and player relevance may be weak", "Define a narrower testable segment"),
    ("Segment criteria missing", f'--({inv_rng("Main gap category")}="Segment criteria missing")', "Targeting cannot be reproduced", "Document actual selection rules"),
    ("Success measure missing", f'--({inv_rng("Main gap category")}="Success measure missing")', "No agreed definition of success", "Agree metric and target before launch"),
    ("Campaign owner unclear", f'--({inv_rng("Campaign owner")}="")', "Decisions and follow-up may stall", "Assign accountable campaign owner"),
    ("Data owner unclear", f'--({inv_rng("Data owner")}="")', "Result data cannot be requested efficiently", "Confirm data owner and access path"),
    ("Result data missing", f'--((({inv_rng("Result data available")}="Pending")+({inv_rng("Result data available")}="Not Available")+({inv_rng("Result data available")}="Unknown"))>0)', "Performance analysis cannot begin", "Request claims, cost and deposit outputs"),
    ("Duplicate or overlapping mechanic", f'--({inv_rng("Duplicate or overlap risk")}="High")', "Players may receive conflicting offers", "Compare and merge similar campaigns"),
    ("High cap requires review", f'--({inv_rng("Mechanic gap")}="High cap requires review")', "Bonus exposure may exceed player need", "Review cap utilisation before changing"),
    ("No review date", f'--({inv_rng("Last reviewed date")}="")', "Legacy campaigns may continue without purpose", "Assign a review date and owner"),
    ("No exclusion criteria", f'--((({inv_rng("Exclusion criteria")}="")+({inv_rng("Exclusion criteria")}="Not documented"))>0)', "Abuse and offer overlap risk may increase", "Define exclusions and abuse controls"),
    ("Campaign purpose no longer clear", f'--({inv_rng("Main gap category")}="Campaign purpose no longer clear")', "Legacy cost may continue without clear value", "Confirm purpose or consider sunset"),
]
example_map = {
    "Objective unclear": "SAMPLE – Standard Recurring Bonus", "Objective too general": "SAMPLE – Payday Mega Bonus", "Segment too broad": "SAMPLE – Payday Mega Bonus",
    "Segment criteria missing": "SAMPLE – Welcome Back Players", "Success measure missing": "SAMPLE – VM Weekly Reward", "Campaign owner unclear": "SAMPLE – Standard Recurring Bonus",
    "Data owner unclear": "SAMPLE – Weekend Reload B", "Result data missing": "SAMPLE – New Depositor Follow-Up", "Duplicate or overlapping mechanic": "SAMPLE – Weekend Reload A / B",
    "High cap requires review": "SAMPLE – Payday Mega Bonus", "No review date": "SAMPLE – Standard Recurring Bonus", "No exclusion criteria": "SAMPLE – VM Weekly Reward",
    "Campaign purpose no longer clear": "SAMPLE – Standard Recurring Bonus",
}
for rr, (display, gap_condition, impact, response) in enumerate(gap_categories, start=13):
    gap.cell(rr, 1, display)
    gap.cell(rr, 2, filtered_count([gap_condition]))
    gap.cell(rr, 3, f'=IFERROR(B{rr}/$A$6,0)')
    gap.cell(rr, 4, example_map[display])
    gap.cell(rr, 5, impact)
    gap.cell(rr, 6, response)
    gap.cell(rr, 3).number_format = "0%"
style_table_body(gap, 13, 25, 1, 6, calculated_cols=[2,3])
add_table(gap, "A12:F25", "tblGapSummary")
for col, width in {"A":30,"B":18,"C":20,"D":32,"E":34,"F":34}.items(): gap.column_dimensions[col].width = width

section(gap, 11, "PRIORITY CAMPAIGNS", 8, 18)
priority_headers = ["Campaign ID", "Campaign name", "Campaign category", "Main gap", "Severity", "Business impact", "Required clarification", "PIC", "Due date", "Status"]
for i, h in enumerate(priority_headers, start=8): gap.cell(12, i, h)
style_header_row(gap, 12, 17)
# Hidden helper ranks qualifying campaigns and respects the six management filters.
priority_rank_col = 39  # AM
gap.cell(2, priority_rank_col, "Priority rank")
for helper_row, inv_row in enumerate(range(7, MAX_DATA_ROW + 1), start=3):
    qualifying = f'OR(\'Campaign Inventory\'!{letter("Gap severity")}{inv_row}="Critical",\'Campaign Inventory\'!{letter("Gap severity")}{inv_row}="High",\'Campaign Inventory\'!{letter("Stated campaign objective")}{inv_row}="",\'Campaign Inventory\'!{letter("Stated target segment")}{inv_row}="",\'Campaign Inventory\'!{letter("Result data available")}{inv_row}="Pending",\'Campaign Inventory\'!{letter("Result data available")}{inv_row}="Not Available",\'Campaign Inventory\'!{letter("Result data available")}{inv_row}="Unknown",\'Campaign Inventory\'!{letter("Duplicate or overlap risk")}{inv_row}="High",\'Campaign Inventory\'!{letter("Campaign owner")}{inv_row}="")'
    filter_checks = [
        f'OR($B$3="All",\'Campaign Inventory\'!{letter("Brand")}{inv_row}=$B$3)',
        f'OR($D$3="All",\'Campaign Inventory\'!{letter("Campaign category")}{inv_row}=$D$3)',
        f'OR($F$3="All",\'Campaign Inventory\'!{letter("Campaign source")}{inv_row}=$F$3)',
        f'OR($H$3="All",\'Campaign Inventory\'!{letter("Campaign owner")}{inv_row}=$H$3)',
        f'OR($J$3="All",\'Campaign Inventory\'!{letter("Gap severity")}{inv_row}=$J$3)',
        f'OR($L$3="All",\'Campaign Inventory\'!{letter("Campaign status")}{inv_row}=$L$3)',
    ]
    gap.cell(helper_row, priority_rank_col, f'=IF(AND(\'Campaign Inventory\'!{letter("Campaign ID")}{inv_row}<>"",{qualifying},{",".join(filter_checks)}),COUNT($AM$2:AM{helper_row-1})+1,"")')

priority_source_map = {
    8: "Campaign ID", 9: "Campaign name", 10: "Campaign category", 11: "Main gap category", 12: "Gap severity",
    13: "Business reason for campaign", 14: "Recommended next action", 15: "Promo PIC", 17: "Clarification status",
}
for offset in range(13, 25):
    rank_match = f'MATCH(ROW()-12,$AM$3:$AM${MAX_DATA_ROW - 4},0)'
    for output_col, source_header in priority_source_map.items():
        source_letter = letter(source_header)
        gap.cell(offset, output_col, f'=IFERROR(INDEX(\'Campaign Inventory\'!${source_letter}$7:${source_letter}${MAX_DATA_ROW},{rank_match}),"")')
    gap.cell(offset, 16, "")
    gap.cell(offset, 16).number_format = "yyyy-mm-dd"
style_table_body(gap, 13, 24, 8, 17, calculated_cols=[8,9,10,11,12,13,14,15,17], external_cols=[16])
add_table(gap, "H12:Q24", "tblPriorityCampaigns")
for col in range(8, 18): gap.column_dimensions[get_column_letter(col)].width = 18
gap.column_dimensions["I"].width = 28
gap.column_dimensions["M"].width = 28
gap.column_dimensions["N"].width = 24

# Chart helper tables at AA:AD and charts, formula-driven
chart_specs = [
    ("Campaigns by category", "CampaignCategories", "Campaign category"),
    ("Objective clarity", "ObjectiveClarityStatuses", "Objective clarity"),
    ("Segment clarity", "SegmentClarityStatuses", "Segment clarity"),
    ("Measurement readiness", "MeasurementStatuses", "Measurement readiness"),
]
helper_start_cols = [27, 30, 33, 36]
chart_positions = ["A28", "J28", "A45", "J45"]
for (title, list_name, inv_header), start_col, pos in zip(chart_specs, helper_start_cols, chart_positions):
    vals = list_values[list_name]
    gap.cell(1, start_col, title)
    gap.cell(2, start_col, "Category")
    gap.cell(2, start_col + 1, "Campaigns")
    for rr, val in enumerate(vals, start=3):
        gap.cell(rr, start_col, val)
        gap.cell(rr, start_col + 1, filtered_count([f'--({inv_rng(inv_header)}="{val}")']))
    chart = BarChart()
    chart.type = "bar"
    chart.style = 10
    chart.title = title
    chart.y_axis.title = ""
    chart.x_axis.title = "Campaign count"
    data = Reference(gap, min_col=start_col + 1, min_row=2, max_row=2 + len(vals))
    cats = Reference(gap, min_col=start_col, min_row=3, max_row=2 + len(vals))
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    chart.height = 7.2
    chart.width = 14
    chart.legend = None
    chart.graphical_properties = None
    gap.add_chart(chart, pos)
for c in range(27, 40): gap.column_dimensions[get_column_letter(c)].hidden = True
gap.freeze_panes = "A3"
gap.sheet_view.zoomScale = 80

# -----------------------------------------------------------------------------
# Segment Map
# -----------------------------------------------------------------------------
seg_headers = ["Segment ID", "Segment name", "Segment description", "Player activity status", "Inactive-day range", "Deposit range", "VIP tier", "Primary product", "Secondary product", "Player behaviour", "Exclusion criteria", "Campaigns using this segment", "Number of campaigns", "Campaign categories used", "Main objective", "Segment owner", "Data source", "Segment currently available", "Segment clarity", "Overlap risk", "Missing segment opportunity", "Recommended action", "Notes"]
seg_col = {h:i+1 for i,h in enumerate(seg_headers)}
set_title(segment, "WS1 Segment Map", "Map existing target groups, identify overlap, and find player segments with no clear campaign coverage.", len(seg_headers))
for i,h in enumerate(seg_headers, start=1): segment.cell(5,i,h)
style_header_row(segment,5,len(seg_headers))
style_table_body(segment,6,105,1,len(seg_headers),calculated_cols=[seg_col["Number of campaigns"]],external_cols=[seg_col["Segment owner"],seg_col["Data source"]])
segment_samples = [
    ["SEG-001","Active low-value players","Active players with lower average deposits","Active","","Below RM100","Non-VIP","Multi-Product","","Frequent low-value deposits","Exclude VIP Gold+ and abuse flags","SAMPLE – Active Player Weekly Reload",None,"Retention","Maintain activity","Promotion","CK Analytics CRM","Yes","Clear","Low","Test smaller effective bonus","Run Controlled Test",SAMPLE_LABEL],
    ["SEG-002","Active high-value players","Active players with higher deposit value","Active","","RM500+","Multiple","Multi-Product","","High-value active play","Exclude self-excluded and risk flags","",None,"","Retention / VIP progression","VIP","VIP","Partial","Partly Clear","Low","No dedicated non-VIP high-value campaign","Define Segment Criteria",SAMPLE_LABEL],
    ["SEG-003","Inactive 7–14 days","Recently inactive players","Inactive","7–14 days","Any","Multiple","Multi-Product","","Previously active depositor","Exclude active bonus holders","",None,"","Early reactivation","Sales","CK Analytics CRM","Partial","Clear","Low","Potential early-warning reactivation test","Run Controlled Test",SAMPLE_LABEL],
    ["SEG-004","Inactive 15–30 days","Players inactive for 15 to 30 days","Inactive","15–30 days","Prior depositor","Multiple","Multi-Product","","No login or deposit 15–30 days","Exclude abuse and self-excluded","SAMPLE – 15–30 Day Reactivation",None,"Reactivation","Reactivation","Sales","CK Analytics CRM","Yes","Clear","Low","Measure 72h and 14d outcomes","Prepare Performance Analysis",SAMPLE_LABEL],
    ["SEG-005","Inactive over 30 days","Longer-term lapsed players","Lapsed","31+ days","Prior depositor","Multiple","Multi-Product","","No activity over 30 days","Exclude self-excluded","SAMPLE – Welcome Back Players",None,"Reactivation","Win-back","Stakeholder","Unknown","Partial","Unclear","Medium","Define inactivity bands before targeting","Define Segment Criteria",SAMPLE_LABEL],
    ["SEG-006","Sports-only players","Active Sports players with no Casino activity","Active","","Sports deposit RM100+","Multiple","Sports","Casino","No Casino activity 90 days","Exclude Casino-active players","SAMPLE – Sports to Casino Bridge",None,"Cross-Product","Cross-product conversion","Sales","CK Analytics CRM","Yes","Clear","Low","Test alternative Casino mechanics","Continue Monitoring",SAMPLE_LABEL],
    ["SEG-007","Casino-only players","Active Casino players with no Sports activity","Active","","Any","Multiple","Casino","Sports","No Sports activity 90 days","Exclude Sports-active players","",None,"","Cross-product conversion","Sales","CK Analytics CRM","Partial","Partly Clear","Low","No Sports bridge campaign identified","Define Segment Criteria",SAMPLE_LABEL],
    ["SEG-008","Gold-tier players","Gold players near progression threshold","Active","","Within 20% of target","Gold","Multi-Product","","Near Platinum milestone","Exclude abuse flags","SAMPLE – Gold to Platinum Milestone",None,"VIP Progression","VIP progression","VIP","VIP","Yes","Clear","Low","Confirm progression success target","Confirm Objective",SAMPLE_LABEL],
    ["SEG-009","Platinum-tier players","Existing Platinum players","Active","","High value","Platinum","Multi-Product","","Platinum activity","VIP-managed exclusions","",None,"","Retention","VIP","VIP","Partial","Partly Clear","Low","No dedicated Platinum retention review","Continue Monitoring",SAMPLE_LABEL],
    ["SEG-010","New depositors","Players with exactly one deposit","New","0–7 days","One deposit","Multiple","Multi-Product","","No repeat deposit","Exclude duplicates and abuse","SAMPLE – New Depositor Follow-Up",None,"Retention","Second deposit","Promotion","CK Analytics CRM","Yes","Clear","Low","Analytics result extract missing","Request Result Data",SAMPLE_LABEL],
    ["SEG-011","Lapsed depositors","Past depositors currently inactive","Lapsed","30+ days","Prior depositor","Multiple","Multi-Product","","No recent deposit","Exclude self-excluded","SAMPLE – Welcome Back Players",None,"Reactivation","Win-back","Sales","Unknown","Partial","Unclear","Medium","Split by lapse duration and past value","Define Segment Criteria",SAMPLE_LABEL],
    ["SEG-012","High-frequency low-value claimants","Players with repeated low-value bonus claims","Active","","Low deposits","Non-VIP","Multi-Product","","Frequent claims with low post-claim value","Exclude from high-cost offers","SAMPLE – Weekend Reload A; SAMPLE – Weekend Reload B",None,"Retention; General Promotion","Cost control","Promotion","CK Analytics CRM","Partial","Partly Clear","High","Possible bonus-overlap and abuse segment","Review Mechanic",SAMPLE_LABEL],
]
for rr, row in enumerate(segment_samples, start=6):
    for cc, value in enumerate(row, start=1): segment.cell(rr,cc,value)
    segment.cell(rr, seg_col["Number of campaigns"], f'=IF(L{rr}="",0,LEN(L{rr})-LEN(SUBSTITUTE(L{rr},";",""))+1)')
for h,named in {"Player activity status":"PlayerActivityStatuses","VIP tier":"VIPTiers","Primary product":"Products","Secondary product":"Products","Segment owner":"OwnerGroups","Segment currently available":"YesNoUnknown","Segment clarity":"SegmentClarityStatuses","Overlap risk":"OverlapRisks","Recommended action":"RecommendedActions"}.items():
    l=get_column_letter(seg_col[h]); add_list_validation(segment,f"{l}6:{l}105",named)
add_table(segment, f"A5:{get_column_letter(len(seg_headers))}105", "tblSegmentMap")
segment.freeze_panes="D6"; segment.sheet_view.zoomScale=75
for i,h in enumerate(seg_headers,1): segment.column_dimensions[get_column_letter(i)].width=16 if h not in ["Segment name","Segment description","Player behaviour","Exclusion criteria","Campaigns using this segment","Missing segment opportunity","Recommended action","Notes"] else 25
segment.conditional_formatting.add(f"{get_column_letter(seg_col['Number of campaigns'])}6:{get_column_letter(seg_col['Number of campaigns'])}105", CellIsRule(operator="greaterThan", formula=[3], fill=PatternFill("solid", fgColor=AMBER)))
segment.conditional_formatting.add(f"{get_column_letter(seg_col['Number of campaigns'])}6:{get_column_letter(seg_col['Number of campaigns'])}105", CellIsRule(operator="equal", formula=[0], fill=PatternFill("solid", fgColor=RED)))
segment.conditional_formatting.add(f"{get_column_letter(seg_col['Segment clarity'])}6:{get_column_letter(seg_col['Segment clarity'])}105", FormulaRule(formula=[f'OR({get_column_letter(seg_col["Segment clarity"])}6="Unclear",{get_column_letter(seg_col["Segment clarity"])}6="Not Provided")'], fill=PatternFill("solid", fgColor=RED)))
segment.conditional_formatting.add(f"{get_column_letter(seg_col['Overlap risk'])}6:{get_column_letter(seg_col['Overlap risk'])}105", FormulaRule(formula=[f'{get_column_letter(seg_col["Overlap risk"])}6="High"'], fill=PatternFill("solid", fgColor=AMBER)))

# -----------------------------------------------------------------------------
# Data Source Map
# -----------------------------------------------------------------------------
ds_headers=["Data field","Business definition","Why it is needed","Primary owner","Supporting owner","Likely system or source","Available now","Access available","Extraction method","Update frequency","Data quality","Current status","Request date","Request sent to","Expected response date","Blocker","Next action","Notes"]
ds_col={h:i+1 for i,h in enumerate(ds_headers)}
set_title(data_map,"WS1 Data Source Map","Map required fields, ownership, access, extraction method and blockers before requesting full performance data.",len(ds_headers))
# summary cards
data_fields=["Campaign objective","Campaign target segment","Campaign dates","Campaign owner","Targeted-player count","Eligible-player count","Bonus mechanic","Minimum deposit","Maximum bonus cap","Turnover requirement","Campaign budget","Promo cost","Total claims","Unique claimants","Deposit amount","NGR","GGR","Reactivated players","Cross-product converted players","Repeat deposit behaviour","Player activity after campaign","Abuse or duplicate-claim flags","VM weekly campaign details","Stakeholder campaign brief","CRM campaign segment","Campaign success target"]
cards=[("A3","A4","Total required data fields",f'=COUNTA(A8:A{7+len(data_fields)})'),("C3","C4","Available fields",f'=COUNTIF(L8:L{7+len(data_fields)},"Available")'),("E3","E4","Partial fields",f'=COUNTIF(L8:L{7+len(data_fields)},"Partial")'),("G3","G4","Pending fields",f'=COUNTIF(L8:L{7+len(data_fields)},"Pending")'),("I3","I4","Unclear ownership",f'=COUNTIF(D8:D{7+len(data_fields)},"Unknown")'),("K3","K4","Blocked by access",f'=COUNTIF(L8:L{7+len(data_fields)},"Blocked")'),("M3","M4","Data readiness %",f'=IFERROR((COUNTIF(L8:L{7+len(data_fields)},"Available")+0.5*COUNTIF(L8:L{7+len(data_fields)},"Partial"))/COUNTA(A8:A{7+len(data_fields)}),0)')]
for lab,val,label_text,formula in cards: card(data_map,lab,val,label_text,formula,"0%" if "%" in label_text else "0")
for i,h in enumerate(ds_headers,1): data_map.cell(7,i,h)
style_header_row(data_map,7,len(ds_headers)); style_table_body(data_map,8,107,1,len(ds_headers),external_cols=[ds_col["Primary owner"],ds_col["Supporting owner"],ds_col["Request sent to"]])
defs={
"Campaign objective":("The intended business or player outcome","Needed to judge whether the campaign worked","Stakeholder","Promotion","Campaign brief"),
"Campaign target segment":("The intended player group","Needed to assess targeting relevance","Sales","Promotion","Campaign brief / CRM"),
"Campaign dates":("Start and end dates","Needed to define the measurement window","Promotion","Sales","Promo setup / calendar"),
"Campaign owner":("Accountable business owner","Needed for decisions and clarification","Promotion","Stakeholder","Campaign brief"),
"Targeted-player count":("Players selected for communication","Needed for targeting and reach rates","CK Analytics","Sales","CRM export"),
"Eligible-player count":("Players meeting mechanic rules","Needed to separate targeting from eligibility","Promotion","CK Analytics","Promo platform"),
"Bonus mechanic":("Reward type and calculation","Needed to compare campaign structures","Promotion","Sales","Promo setup"),
"Minimum deposit":("Minimum qualifying deposit","Needed for mechanic and cost review","Promotion","Finance","Promo setup"),
"Maximum bonus cap":("Maximum reward per player","Needed for cap and cost exposure review","Promotion","Finance","Promo setup"),
"Turnover requirement":("Wagering requirement","Needed for mechanic and player-value review","Promotion","Management","Promo setup"),
"Campaign budget":("Approved total campaign budget","Needed for cost governance","Management","Finance","Approval record"),
"Promo cost":("Actual bonus cost","Needed for later efficiency analysis","Finance","CK Analytics","Finance / promo ledger"),
"Total claims":("Number of claims","Needed for campaign response measurement","Promotion","CK Analytics","Promo platform"),
"Unique claimants":("Distinct players claiming","Needed to avoid duplicate claim counts","CK Analytics","Promotion","Analytics extract"),
"Deposit amount":("Deposits in agreed campaign window","Needed for directional response analysis","CK Analytics","Finance","Transaction data"),
"NGR":("Net gaming revenue","Needed only for validated future profitability analysis","CK Analytics","Finance","Analytics warehouse"),
"GGR":("Gross gaming revenue","Needed for future revenue context","CK Analytics","Finance","Analytics warehouse"),
"Reactivated players":("Previously inactive players returning to deposit","Needed for reactivation measurement","CK Analytics","Sales","CRM / transaction data"),
"Cross-product converted players":("Players adopting the target product","Needed for cross-product measurement","CK Analytics","Sales","Product activity data"),
"Repeat deposit behaviour":("Deposits after the initial campaign action","Needed for retention context","CK Analytics","Sales","Transaction data"),
"Player activity after campaign":("Post-campaign play and activity","Needed to understand durability","CK Analytics","Promotion","Analytics warehouse"),
"Abuse or duplicate-claim flags":("Player or claim risk indicators","Needed for cost and control review","Promotion","CK Analytics","Risk / promo platform"),
"VM weekly campaign details":("Weekly VM list, mechanic and timing","Needed to document recurring retention work","VIP","Promotion","VM tracker"),
"Stakeholder campaign brief":("Original request and rationale","Needed to confirm campaign intent","Stakeholder","Promotion","Brief / email"),
"CRM campaign segment":("Actual CRM selection logic","Needed to reproduce target segments","CK Analytics","Sales","CRM"),
"Campaign success target":("Planned performance target","Needed to define success before launch","Stakeholder","Sales","Campaign brief"),}
for rr, field in enumerate(data_fields,start=8):
    definition,why,owner,support,system=defs[field]
    status = "Available" if field in ["Campaign dates","Bonus mechanic","Minimum deposit","Maximum bonus cap","Turnover requirement","Total claims"] else ("Partial" if field in ["Campaign objective","Campaign target segment","Campaign owner","Targeted-player count","Eligible-player count","Promo cost","Unique claimants","VM weekly campaign details","Stakeholder campaign brief"] else "Pending")
    row=[field,definition,why,owner,support,system,"Yes" if status=="Available" else "Partial" if status=="Partial" else "Unknown","Yes" if status=="Available" else "Partial","Manual export" if "brief" not in system.lower() else "Document review","Weekly" if field in ["Total claims","Promo cost","Deposit amount"] else "Per campaign","Acceptable" if status!="Pending" else "Unknown",status,"","","","","Confirm owner and request sample extract" if status=="Pending" else "Validate definition and field format",SAMPLE_LABEL]
    for cc,val in enumerate(row,1): data_map.cell(rr,cc,val)
for h,named in {"Primary owner":"OwnerGroups","Supporting owner":"OwnerGroups","Available now":"YesNoUnknown","Access available":"AccessStatuses","Data quality":"DataQualityStatuses","Current status":"DataMapStatuses"}.items():
    l=get_column_letter(ds_col[h]); add_list_validation(data_map,f"{l}8:{l}107",named)
for h in ["Request date","Expected response date"]:
    for rr in range(8,108): data_map.cell(rr,ds_col[h]).number_format="yyyy-mm-dd"
add_table(data_map,f"A7:{get_column_letter(len(ds_headers))}107","tblDataSourceMap")
data_map.freeze_panes="D8"; data_map.sheet_view.zoomScale=75
for i,h in enumerate(ds_headers,1): data_map.column_dimensions[get_column_letter(i)].width=16 if h not in ["Business definition","Why it is needed","Likely system or source","Blocker","Next action","Notes"] else 25
data_map["M4"].number_format="0%"
data_map.conditional_formatting.add(f"L8:L107", FormulaRule(formula=['L8="Blocked"'], fill=PatternFill("solid",fgColor=RED)))
data_map.conditional_formatting.add(f"D8:D107", FormulaRule(formula=['D8="Unknown"'], fill=PatternFill("solid",fgColor=AMBER)))

# -----------------------------------------------------------------------------
# Action Tracker
# -----------------------------------------------------------------------------
act_headers=["Action ID","Campaign ID","Campaign name","Gap category","Issue","Required action","Information required","Request to","PIC","Priority","Start date","Due date","Status","Response received","Resolution","Follow-up required","Last updated","Notes"]
act_col={h:i+1 for i,h in enumerate(act_headers)}
set_title(actions,"WS1 Clarification Action Tracker","Turn discovery gaps into owned requests, due dates and documented resolutions.",len(act_headers))
cards=[("A3","A4","Total open actions",'=COUNTIFS(M8:M107,"<>Completed",A8:A107,"<>")'),("C3","C4","High-priority actions",'=COUNTIFS(J8:J107,"High",M8:M107,"<>Completed")+COUNTIFS(J8:J107,"Critical",M8:M107,"<>Completed")'),("E3","E4","Overdue actions",'=COUNTIFS(L8:L107,"<"&TODAY(),L8:L107,"<>",M8:M107,"<>Completed",A8:A107,"<>")'),("G3","G4","Waiting for stakeholder",'=COUNTIFS(H8:H107,"Stakeholder",M8:M107,"Waiting for Response")'),("I3","I4","Waiting for Analytics",'=COUNTIFS(H8:H107,"CK Analytics",M8:M107,"Waiting for Response")'),("K3","K4","Completed actions",'=COUNTIF(M8:M107,"Completed")')]
for lab,val,label_text,formula in cards: card(actions,lab,val,label_text,formula)
for i,h in enumerate(act_headers,1): actions.cell(7,i,h)
style_header_row(actions,7,len(act_headers)); style_table_body(actions,8,107,1,len(act_headers),external_cols=[act_col["Request to"],act_col["Response received"]])
action_samples=[
["ACT-SMP-001","WS1-SMP-004","SAMPLE – Welcome Back Players","Segment criteria missing","Inactive-day rule is not defined","Define inactive-day bands and eligibility rules","Confirmed inactivity window and prior-deposit rule","Stakeholder","Promo PIC B","Critical",date(2026,7,22),date(2026,7,25),"Not Started","No","","Yes",date(2026,7,22),SAMPLE_LABEL],
["ACT-SMP-002","WS1-SMP-008","SAMPLE – Standard Recurring Bonus","Campaign purpose no longer clear","Objective and owner are not documented","Confirm purpose and accountable owner","Original rationale, current owner and review decision","Management","Promo PIC A","Critical",date(2026,7,22),date(2026,7,29),"Not Started","No","","Yes",date(2026,7,22),SAMPLE_LABEL],
["ACT-SMP-003","WS1-SMP-009","SAMPLE – Weekend Reload A","Duplicate or overlapping mechanic","Similar players and mechanic overlap with B","Compare both campaigns and decide whether to merge","Target lists, mechanics, timing and cost","Sales","Promo PIC B","High",date(2026,7,22),date(2026,7,31),"In Progress","Partial","","Yes",date(2026,7,22),SAMPLE_LABEL],
["ACT-SMP-004","WS1-SMP-010","SAMPLE – Weekend Reload B","Data owner unclear","No data owner confirmed","Confirm result-data owner","Named owner and available result fields","Stakeholder","Promo PIC C","High",date(2026,7,22),date(2026,7,26),"Waiting for Response","No","","Yes",date(2026,7,22),SAMPLE_LABEL],
["ACT-SMP-005","WS1-SMP-011","SAMPLE – New Depositor Follow-Up","Result data missing","Objectives are complete but Analytics output is missing","Request campaign result extract","Claims, promo cost, repeat deposit and dates","CK Analytics","Promo PIC A","High",date(2026,7,22),date(2026,7,30),"Waiting for Response","No","","Yes",date(2026,7,22),SAMPLE_LABEL],
["ACT-SMP-006","WS1-SMP-007","SAMPLE – Payday Mega Bonus","Segment too broad","All contactable players are targeted","Propose segmented controlled test","Player segments, sample sizes and exclusions","Management","Promo PIC B","Critical",date(2026,7,22),date(2026,7,28),"In Progress","Partial","","Yes",date(2026,7,22),SAMPLE_LABEL],
["ACT-SMP-007","WS1-SMP-007","SAMPLE – Payday Mega Bonus","High cap requires review","RM1,000 cap may create excess exposure","Review cap usage before changing","Claim deposits, actual bonuses and cap utilisation","CK Analytics","Promo PIC B","High",date(2026,7,22),date(2026,8,3),"Not Started","No","","Yes",date(2026,7,22),SAMPLE_LABEL],
["ACT-SMP-008","WS1-SMP-002","SAMPLE – VM Weekly Reward","Success measure missing","Planned target is not documented","Agree a practical weekly retention target","VM objective, target and available outputs","VIP","Promo PIC B","Medium",date(2026,7,22),date(2026,7,29),"Waiting for Response","No","","Yes",date(2026,7,22),SAMPLE_LABEL],
]
for rr,row in enumerate(action_samples,start=8):
    for cc,val in enumerate(row,1): actions.cell(rr,cc,val)
for h,named in {"Request to":"OwnerGroups","Priority":"Priorities","Status":"ActionStatuses","Response received":"YesNoUnknown","Follow-up required":"YesNoUnknown"}.items():
    l=get_column_letter(act_col[h]); add_list_validation(actions,f"{l}8:{l}107",named)
for h in ["Start date","Due date","Last updated"]:
    for rr in range(8,108): actions.cell(rr,act_col[h]).number_format="yyyy-mm-dd"
add_table(actions,f"A7:{get_column_letter(len(act_headers))}107","tblActionTracker")
actions.freeze_panes="D8"; actions.sheet_view.zoomScale=80
for i,h in enumerate(act_headers,1): actions.column_dimensions[get_column_letter(i)].width=16 if h not in ["Campaign name","Issue","Required action","Information required","Resolution","Notes"] else 25
actions.conditional_formatting.add("A8:R107", FormulaRule(formula=['AND($A8<>"",$L8<TODAY(),$L8<>"",$M8<>"Completed")'], fill=PatternFill("solid",fgColor=RED)))
actions.conditional_formatting.add("J8:J107", FormulaRule(formula=['J8="Critical"'], fill=PatternFill("solid",fgColor=RED), font=Font(color=RED_TEXT,bold=True)))
actions.conditional_formatting.add("M8:M107", FormulaRule(formula=['M8="Waiting for Response"'], fill=PatternFill("solid",fgColor=AMBER)))
actions.conditional_formatting.add("M8:M107", FormulaRule(formula=['M8="Completed"'], fill=PatternFill("solid",fgColor=GREEN)))

# Final workbook configuration
lists.sheet_state = "hidden"
wb.calculation.fullCalcOnLoad = True
wb.calculation.forceFullCalc = True
wb.calculation.calcMode = "auto"

# Print settings for management sheets
for ws in [readme, gap, data_map, actions]:
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.page_margins.left = 0.25
    ws.page_margins.right = 0.25
    ws.page_margins.top = 0.5
    ws.page_margins.bottom = 0.5

wb.save(OUT)

# -----------------------------------------------------------------------------
# Programmatic validation after save
# -----------------------------------------------------------------------------
check = load_workbook(OUT, data_only=False)
assert check.sheetnames == sheet_names, check.sheetnames
assert check["Lists"].sheet_state == "hidden"
assert "tblCampaignInventory" in check["Campaign Inventory"].tables
assert "tblSegmentMap" in check["Segment Map"].tables
assert "tblDataSourceMap" in check["Data Source Map"].tables
assert "tblActionTracker" in check["Action Tracker"].tables
assert len(check["Campaign Inventory"].data_validations.dataValidation) >= 10
assert len(check["Action Tracker"].data_validations.dataValidation) >= 3
sample_names = [check["Campaign Inventory"].cell(r, inv_col["Campaign name"]).value for r in range(7, 7 + len(samples))]
assert len(samples) >= 10
assert all("SAMPLE" in name for name in sample_names)
assert all(check["Campaign Inventory"].cell(r, inv_col["Notes"]).value == SAMPLE_LABEL for r in range(7, 7 + len(samples)))

formula_cells = []
for ws in check.worksheets:
    for row in ws.iter_rows():
        for cell in row:
            if isinstance(cell.value, str) and cell.value.startswith("="):
                formula_cells.append((ws.title, cell.coordinate, cell.value))
                assert "#REF!" not in cell.value and "#DIV/0!" not in cell.value

summary = {
    "final_file_path": str(OUT),
    "tabs_created": check.sheetnames,
    "sample_campaigns": len(samples),
    "main_formulas": ["Objective completeness score", "Segment completeness score", "Measurement readiness score", "Overall readiness", "Gap KPIs", "Data readiness", "Action KPIs"],
    "main_dropdown_lists": ["Campaign source", "Campaign category", "Campaign status", "Frequency", "Objective clarity", "Segment clarity", "Data availability", "Gap severity", "Recommended action", "Delivery channel", "QC status", "Priority", "Action status"],
    "tables": {ws.title: list(ws.tables.keys()) for ws in check.worksheets if ws.tables},
    "formula_count": len(formula_cells),
    "google_sheets_limitations": ["Excel table styling, named-range dropdowns and chart layout may render slightly differently after Google Sheets import.", "Formulas use common Excel/Google Sheets functions and avoid macros, pivots and Power Query."],
}
print(json.dumps(summary, indent=2, default=str))
