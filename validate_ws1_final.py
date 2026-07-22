import json
import re
import zipfile
from pathlib import Path

from openpyxl import load_workbook


path = Path(r"C:\Users\vdiuser\Downloads\promo-automation\WS1_Campaign_Discovery_and_Gap_Analysis_v0.1.xlsx")
expected_sheets = [
    "README",
    "Campaign Inventory",
    "Gap Summary",
    "Segment Map",
    "Data Source Map",
    "Action Tracker",
    "Lists",
]
required_tables = {
    "Campaign Inventory": "tblCampaignInventory",
    "Gap Summary": "tblGapSummary",
    "Segment Map": "tblSegmentMap",
    "Data Source Map": "tblDataSourceMap",
    "Action Tracker": "tblActionTracker",
}

assert path.exists() and path.stat().st_size > 50_000
with zipfile.ZipFile(path) as archive:
    assert archive.testzip() is None
    names = set(archive.namelist())
    assert "xl/workbook.xml" in names
    assert not any("vbaProject" in name for name in names)

wb = load_workbook(path, data_only=False)
assert wb.sheetnames == expected_sheets
assert wb["Lists"].sheet_state == "hidden"
assert not wb._external_links

for sheet_name, table_name in required_tables.items():
    assert table_name in wb[sheet_name].tables
assert "tblPriorityCampaigns" in wb["Gap Summary"].tables
assert sum(len(ws._charts) for ws in wb.worksheets) == 4

assert wb["Campaign Inventory"].freeze_panes == "H7"
assert wb["Gap Summary"].freeze_panes == "A3"
assert wb["Segment Map"].freeze_panes == "D6"
assert wb["Data Source Map"].freeze_panes == "D8"
assert wb["Action Tracker"].freeze_panes == "D8"

named_ranges = set(wb.defined_names.keys())
for required_name in [
    "Brands", "CampaignSources", "CampaignCategories", "CampaignStatuses",
    "Frequencies", "ObjectiveClarityStatuses", "SegmentClarityStatuses",
    "DataAvailabilityStatuses", "GapSeverities", "ClarificationStatuses",
    "RecommendedActions", "DeliveryChannels", "QCStatuses", "OwnerGroups",
    "Priorities", "ActionStatuses", "BonusTypes", "Currencies", "Products",
    "VIPTiers", "PlayerActivityStatuses", "ReadyThreshold", "PartialThreshold",
]:
    assert required_name in named_ranges

validation_counts = {
    ws.title: len(ws.data_validations.dataValidation)
    for ws in wb.worksheets
}
assert validation_counts["Campaign Inventory"] >= 25
assert validation_counts["Gap Summary"] >= 6
assert validation_counts["Segment Map"] >= 8
assert validation_counts["Data Source Map"] >= 6
assert validation_counts["Action Tracker"] >= 5

inventory = wb["Campaign Inventory"]
headers = {inventory.cell(6, col).value: col for col in range(1, inventory.max_column + 1)}
sample_rows = list(range(7, 19))
sample_names = [inventory.cell(row, headers["Campaign name"]).value for row in sample_rows]
sample_notes = [inventory.cell(row, headers["Notes"]).value for row in sample_rows]
assert len(sample_rows) == 12
assert all(name and "SAMPLE" in name for name in sample_names)
assert all(note == "SAMPLE DATA – REPLACE" for note in sample_notes)
assert all(ord(note[12]) == 0x2013 for note in sample_notes)

scenario_tokens = [
    "Active Player Weekly Reload", "VM Weekly Reward", "15–30 Day Reactivation",
    "Welcome Back Players", "Sports to Casino Bridge", "Gold to Platinum Milestone",
    "Payday Mega Bonus", "Standard Recurring Bonus", "Weekend Reload A",
    "Weekend Reload B", "New Depositor Follow-Up",
]
assert all(any(token in name for name in sample_names) for token in scenario_tokens)

formula_cells = []
bad_formula_cells = []
bad_tokens = re.compile(r"#REF!|#DIV/0!|#VALUE!|#NAME\?|#N/A")
for ws in wb.worksheets:
    for row in ws.iter_rows():
        for cell in row:
            value = cell.value
            if isinstance(value, str) and value.startswith("="):
                formula_cells.append(f"{ws.title}!{cell.coordinate}")
                if bad_tokens.search(value):
                    bad_formula_cells.append(f"{ws.title}!{cell.coordinate}")
assert not bad_formula_cells
assert len(formula_cells) >= 1_900

for field in [
    "Objective completeness score",
    "Segment completeness score",
    "Measurement readiness score",
    "Overall readiness",
]:
    assert isinstance(inventory.cell(7, headers[field]).value, str)
    assert inventory.cell(7, headers[field]).value.startswith("=")
assert "ReadyThreshold" in inventory.cell(7, headers["Overall readiness"]).value
assert "PartialThreshold" in inventory.cell(7, headers["Overall readiness"]).value

readme_statement = wb["README"]["A4"].value
assert readme_statement == "This workbook is for campaign discovery and gap analysis only. It is not yet the final Promotion Performance Report."

summary = {
    "status": "PASS",
    "final_file_path": str(path),
    "file_size_bytes": path.stat().st_size,
    "tabs_created": wb.sheetnames,
    "sample_campaigns": len(sample_names),
    "excel_tables": {
        ws.title: list(ws.tables.keys()) for ws in wb.worksheets if ws.tables
    },
    "dropdown_validation_rules": validation_counts,
    "formula_count": len(formula_cells),
    "chart_count": sum(len(ws._charts) for ws in wb.worksheets),
    "lists_tab_hidden": wb["Lists"].sheet_state == "hidden",
    "macros_present": False,
    "formula_error_tokens_found": 0,
    "google_sheets_limitations": [
        "Excel table styling, named-range dropdowns and chart layout may render slightly differently after Google Sheets import.",
        "No macros, pivots or Power Query are used; formulas rely on common Excel/Google Sheets functions.",
    ],
}
print(json.dumps(summary, indent=2, ensure_ascii=False))
